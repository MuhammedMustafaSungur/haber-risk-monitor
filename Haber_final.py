import asyncio
import os
import re
import time
from datetime import datetime, timedelta, timezone
from urllib.parse import quote
from concurrent.futures import ThreadPoolExecutor, as_completed

import feedparser
import pandas as pd
import requests
from bs4 import BeautifulSoup

try:
    import trafilatura
except Exception:
    trafilatura = None

from rapidfuzz import fuzz
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


# ============================================================
# CONFIG
# ============================================================

DAYS_BACK = 1
# Belirli bir günü test etmek için YYYY-MM-DD gir.
# Günlük otomatik kullanımda None yap: TARGET_DATE = None
TARGET_DATE = "2026-05-14"
MAX_NEWS = 100

LANG = "tr"
COUNTRY = "TR"

USE_MAPS = True
ENABLE_ARTICLE_FETCH = True   # Tam içerik çekilebilir; firma yakalama alan bazlı skorla kontrol edilir.
FETCH_ARTICLE_IF_SUMMARY_SHORT = True
MIN_SUMMARY_LENGTH_FOR_NO_FETCH = 250
ARTICLE_FETCH_TIMEOUT = 10
RSS_WORKERS = 8

# Google News RSS linkleri bazen requests ile gerçek haber sitesine gitmez.
# Bu yüzden kısa summary kalan haberlerde browser fallback gerçek haber sayfasına yönlenmeyi dener.
USE_BROWSER_ARTICLE_FETCH_FALLBACK = True
BROWSER_ARTICLE_FETCH_MAX_ROWS = 999         # v20: aday sayısı skora göre; pratikte limit yok
BROWSER_ARTICLE_WAIT_MS = 1200               # v21: v17 gibi biraz bekle; içerik geç yükleniyordu
BROWSER_ARTICLE_GOTO_TIMEOUT_MS = 9000       # v21: dom yükleme yerine commit ile kullanılıyor
BROWSER_ARTICLE_SINGLE_TIMEOUT_SECONDS = 12  # v21: v17 stabilitesine yakın
BROWSER_ARTICLE_TOTAL_TIME_LIMIT_SECONDS = 180 # v15: tüm fallback adımı için toplam sınır
BROWSER_ARTICLE_CONCURRENCY = 2              # v21: içerik çekme stabilitesi için 2; 3 bazı sitelerde boş DOM döndürüyordu
BROWSER_RESTART_EVERY_N_ARTICLES = 8          # v14: takılmaları azaltmak için periyodik context yenileme
MIN_ARTICLE_TEXT_LENGTH = 250

# v15: Browser fallback herkese gitmez; sadece bilgi eksiği kalan değerli haberlere gider.
FALLBACK_INFO_SCORE_THRESHOLD = 70
FALLBACK_REQUIRE_MISSING_CORE_FIELD = True

# v13: Tam içerik açıkken firma yakalamanın patlamaması için
# aday firmalar başlık/summary/gövde alanlarına göre puanlanır.
COMPANY_MIN_SCORE = 72
COMPANY_MAX_CANDIDATES = 7
COMPANY_BODY_MAX_SCORE = 45
COMPANY_TITLE_WEIGHT = 100
COMPANY_SUMMARY_WEIGHT = 70
COMPANY_BODY_WEIGHT = 25

MAPS_MAX_QUERIES_PER_NEWS = 2
MAPS_MAX_RESULTS = 5

SPONSORED_MAPS_WORDS = [
    "sponsorlu", "sponsored", "reklam", "ad"
]

# v14: firma adayı olarak asla kabul edilmeyecek kamu/kişi/haber/jenerik ifadeler.
# Amaç: Jandarma, Emniyet, elektrikçi, elektronik para gibi gerçek müşteri/firma olmayan adayları kesmek.
COMPANY_HARD_BLOCK_WORDS = [
    "jandarma", "emniyet", "polis", "savcılık", "savcilik", "mahkeme", "valilik",
    "kaymakamlık", "kaymakamlik", "belediye", "bakanlık", "bakanlik", "müdürlüğü", "mudurlugu",
    "başkanlığı", "baskanligi", "komutanlığı", "komutanligi", "ekipleri", "ekibi",
    "şüpheli", "supheli", "sanık", "sanik", "fail", "zanlı", "zanli", "tutuklandı", "tutuklandi",
    "gözaltı", "gozalti", "operasyon", "soruşturma", "sorusturma", "iddianame",
    "elektronik para", "ödeme kuruluşu", "odeme kurulusu", "finansal kuruluş", "finansal kurulus",
    "elektrikçi", "elektrikci", "elektrik ustası", "elektrik ustasi", "oto elektrik",
    "yangın", "yangin", "haber", "gazete", "ajans", "son dakika",
    "asliye", "ticaret mahkemesi", "mahkemesi", "icra dairesi", "cumhuriyet başsavcılığı", "cumhuriyet bassavciligi",
]

COMPANY_GENERIC_PHRASES = [
    "elektronik para kuruluşu", "elektronik para kurulusu",
    "elektronik para ve ödeme kuruluşu", "elektronik para ve odeme kurulusu",
    "ödeme kuruluşu", "odeme kurulusu", "finansal kuruluş", "finansal kurulus",
    "elektrikçi", "elektrikci", "jandarma", "emniyet", "polis",
    "firma", "şirket", "sirket", "işletme", "isletme", "kuruluş", "kurulus",
    "asliye ticaret", "ticaret mahkemesi", "icra dairesi"
]

COMPANY_CONTEXT_BAD_WORDS = [
    "tarafından", "tarafindan", "hakkında", "hakkinda", "nedeniyle", "sonucu",
    "yakalandı", "yakalandi", "gözaltına", "gozaltina", "ele geçirildi", "ele gecirildi",
    "arama", "baskın", "baskin", "operasyon", "jandarma", "polis", "emniyet",
    "asliye", "mahkemesi", "icra dairesi", "cumhuriyet başsavcılığı", "cumhuriyet bassavciligi"
]

OUTPUT_DIR = "data"
CSV_FILE = os.path.join(OUTPUT_DIR, "haberler_final.csv")
XLSX_FILE = os.path.join(OUTPUT_DIR, "haberler_final.xlsx")git --version

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}


# ============================================================
# HELPERS
# ============================================================

def normalize_text(text):
    if not text:
        return ""

    text = str(text).lower()

    replacements = {
        "ı": "i",
        "ğ": "g",
        "ü": "u",
        "ş": "s",
        "ö": "o",
        "ç": "c",
        "İ": "i",
    }

    for k, v in replacements.items():
        text = text.replace(k, v)

    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text)

    return text.strip()


def clean_spaces(text):
    return re.sub(r"\s+", " ", str(text)).strip()


def remove_source_from_title(title):
    title = str(title)
    title = re.sub(r"\s+[-–|]\s+.*$", "", title)
    return clean_spaces(title)


def keyword_match(text, keywords):
    text_norm = normalize_text(text)
    return any(normalize_text(k) in text_norm for k in keywords)


def parse_google_news_date(entry):
    published_parsed = entry.get("published_parsed")
    if published_parsed:
        return datetime(*published_parsed[:6], tzinfo=timezone.utc)
    return None


# ============================================================
# KEYWORDS
# ============================================================

SEARCH_TERMS = [
    "yangın", "fabrika yangını", "depo yangını",
    "tesis yangını", "üretim tesisi yangını",

    "kara para", "para aklama", "MASAK", "MASAK raporu",
    "mal varlığına el konuldu",
    "yasa dışı bahis", "yasadışı bahis", "bahis operasyonu",
    "dolandırıcılık", "nitelikli dolandırıcılık",
    "sahtecilik",

    "iflas", "konkordato", "mali kriz",
    "borç krizi", "ödeme güçlüğü", "tasfiye",

    "usulsüzlük", "yolsuzluk", "rüşvet", "zimmet",
    "vergi kaçakçılığı", "sahte fatura",
    "mali suç", "suç örgütü", "örgütlü suç",
    "soruşturma", "gözaltı", "tutuklama",
    "şirketlere operasyon", "firmalara operasyon",
    "vergi incelemesi", "vergi cezası",
    "ihale yolsuzluğu", "ihaleye fesat karıştırma",

    "banka soruşturma", "bankaya soruşturma",
    "banka dolandırıcılığı", "banka hesapları bloke",
    "banka hesaplarına bloke", "banka hesabına bloke",
    "hesaplara bloke", "şüpheli işlem", "şüpheli para transferi",

    "ödeme kuruluşu", "elektronik para ve ödeme kuruluşu",
    "elektronik para ve ödeme kuruluşları",
    "elektronik para kuruluşu", "lisans iptali", "faaliyet izni iptali",
]

TITLE_KEYWORDS = [
    "yangın", "fabrika yangını", "depo yangını", "tesis yangını",

    "kara para", "para aklama", "aklama", "masak",
    "bahis", "yasa dışı bahis", "yasadışı bahis",
    "dolandırıcılık", "sahtecilik", "resmi belgede sahtecilik",

    "iflas", "konkordato",
    "mali kriz", "borç krizi", "ödeme güçlüğü", "tasfiye",

    "usulsüzlük", "vergi kaçakçılığı",
    "sahte fatura", "rüşvet", "zimmet", "yolsuzluk",
    "mali suç", "suç örgütü", "örgütlü suç",
    "soruşturma", "gözaltı", "tutuklama",
    "mal varlığı", "el konuldu",
    "ihale", "ihaleye fesat",
    "vergi incelemesi", "vergi cezası",

    "banka", "şüpheli işlem",
    "para transferi", "aracı kurum", "aracı kuruluş",
    "yatırım kuruluşu", "portföy yönetim",
    "varlık yönetim", "ödeme kuruluşu",
    "elektronik ödeme kuruluşu",
    "elektronik para kuruluşu",
    "elektronik para ve ödeme kuruluşu",
    "elektronik para ve ödeme kuruluşları",
    "lisans iptali", "faaliyet izni iptali"
]

CONTENT_KEYWORDS = [
    "şirket", "firma", "işletme", "kuruluş",
    "limited", "ltd", "anonim şirket",
    "holding", "grup şirketi", "iş insanı",

    "banka", "bankacılık", "finans kuruluşu",
    "finansal kuruluş", "kredi kuruluşu",
    "ödeme kuruluşu", "elektronik ödeme kuruluşu",
    "elektronik para kuruluşu",
    "elektronik para ve ödeme kuruluşu",
    "elektronik para ve ödeme kuruluşları",

    "banka hesabı", "hesaplara bloke",
    "şüpheli işlem", "para transferi",
    "masak", "mal varlığı", "el konuldu",

    "tesis", "fabrika", "depo", "üretim tesisi",
    "satış tesisi", "satış deposu", "dağıtım merkezi",
    "akaryakıt istasyonu", "istasyon",
    "lojistik", "lojistik deposu", "antrepo", "gümrük",

    "sanayi", "ticaret", "organize sanayi", "osb",
    "üretim", "imalat", "imalathane", "atölye",
    "sağlık", "şantiye",
    "maden", "enerji", "inşaat", "tekstil",
    "toptan ve perakende",
    "otomotiv", "gıda", "turizm", "nakliye"
]

EXCLUDE_NEWS_KEYWORDS = [
    "tatbikat", "yangın tatbikatı", "deprem tatbikatı",
    "iş yeri", "işyeri", "ofis", "ofisi",
    "haciz", "icra", "kaçakçılık"
]


# ============================================================
# LOCATION DICTIONARY
# ============================================================

LOCATION_DICTIONARY = {
    "Adana": {"districts": ["Aladağ", "Ceyhan", "Çukurova", "Feke", "İmamoğlu", "Karaisalı", "Karataş", "Kozan", "Pozantı", "Saimbeyli", "Sarıçam", "Seyhan", "Tufanbeyli", "Yumurtalık", "Yüreğir"], "neighborhoods": ["Adana Hacı Sabancı OSB"]},
    "Adıyaman": {"districts": ["Besni", "Çelikhan", "Gerger", "Gölbaşı", "Kahta", "Merkez", "Samsat", "Sincik", "Tut"], "neighborhoods": []},
    "Afyonkarahisar": {"districts": ["Başmakçı", "Bayat", "Bolvadin", "Çay", "Çobanlar", "Dazkırı", "Dinar", "Emirdağ", "Evciler", "Hocalar", "İhsaniye", "İscehisar", "Kızılören", "Merkez", "Sandıklı", "Sinanpaşa", "Sultandağı", "Şuhut"], "neighborhoods": ["Afyonkarahisar OSB"]},
    "Ağrı": {"districts": ["Diyadin", "Doğubayazıt", "Eleşkirt", "Hamur", "Merkez", "Patnos", "Taşlıçay", "Tutak"], "neighborhoods": []},
    "Aksaray": {"districts": ["Ağaçören", "Eskil", "Gülağaç", "Güzelyurt", "Merkez", "Ortaköy", "Sarıyahşi", "Sultanhanı"], "neighborhoods": ["Aksaray OSB"]},
    "Amasya": {"districts": ["Göynücek", "Gümüşhacıköy", "Hamamözü", "Merkez", "Merzifon", "Suluova", "Taşova"], "neighborhoods": ["Merzifon OSB"]},
    "Ankara": {"districts": ["Altındağ", "Ayaş", "Bala", "Beypazarı", "Çamlıdere", "Çankaya", "Çubuk", "Elmadağ", "Etimesgut", "Evren", "Gölbaşı", "Güdül", "Haymana", "Kalecik", "Kazan", "Keçiören", "Kızılcahamam", "Mamak", "Nallıhan", "Polatlı", "Pursaklar", "Sincan", "Şereflikoçhisar", "Yenimahalle"], "neighborhoods": ["Ostim", "İvedik", "Siteler", "Macunköy", "Balgat", "Başkent OSB"]},
    "Antalya": {"districts": ["Akseki", "Aksu", "Alanya", "Demre", "Döşemealtı", "Elmalı", "Finike", "Gazipaşa", "Gündoğmuş", "İbradı", "Kaş", "Kemer", "Kepez", "Konyaaltı", "Korkuteli", "Kumluca", "Manavgat", "Muratpaşa", "Serik"], "neighborhoods": ["Eski Sanayi", "Antalya OSB", "Altınova", "Güzeloba", "Kızıltoprak", "Çağlayan", "Meltem"]},
    "Ardahan": {"districts": ["Çıldır", "Damal", "Göle", "Hanak", "Merkez", "Posof"], "neighborhoods": []},
    "Artvin": {"districts": ["Ardanuç", "Arhavi", "Borçka", "Hopa", "Kemalpaşa", "Merkez", "Murgul", "Şavşat", "Yusufeli"], "neighborhoods": []},
    "Aydın": {"districts": ["Bozdoğan", "Buharkent", "Çine", "Didim", "Efeler", "Germencik", "İncirliova", "Karacasu", "Karpuzlu", "Koçarlı", "Köşk", "Kuşadası", "Kuyucak", "Nazilli", "Söke", "Sultanhisar", "Yenipazar"], "neighborhoods": ["Aydın OSB", "ASTİM OSB"]},
    "Balıkesir": {"districts": ["Altıeylül", "Ayvalık", "Balya", "Bandırma", "Bigadiç", "Burhaniye", "Dursunbey", "Edremit", "Erdek", "Gönen", "Havran", "İvrindi", "Karesi", "Kepsut", "Manyas", "Marmara", "Savaştepe", "Sındırgı", "Susurluk"], "neighborhoods": ["Bandırma OSB", "Balıkesir OSB"]},
    "Bartın": {"districts": ["Amasra", "Kurucaşile", "Merkez", "Ulus"], "neighborhoods": ["Bartın OSB"]},
    "Batman": {"districts": ["Beşiri", "Gercüş", "Hasankeyf", "Kozluk", "Merkez", "Sason"], "neighborhoods": ["Batman OSB"]},
    "Bayburt": {"districts": ["Aydıntepe", "Demirözü", "Merkez"], "neighborhoods": []},
    "Bilecik": {"districts": ["Bozüyük", "Gölpazarı", "İnhisar", "Merkez", "Osmaneli", "Pazaryeri", "Söğüt", "Yenipazar"], "neighborhoods": ["Bozüyük OSB", "Bilecik OSB"]},
    "Bingöl": {"districts": ["Adaklı", "Genç", "Karlıova", "Kiğı", "Merkez", "Solhan", "Yayladere", "Yedisu"], "neighborhoods": ["Bingöl OSB"]},
    "Bitlis": {"districts": ["Adilcevaz", "Ahlat", "Güroymak", "Hizan", "Merkez", "Mutki", "Tatvan"], "neighborhoods": ["Tatvan OSB"]},
    "Bolu": {"districts": ["Dörtdivan", "Gerede", "Göynük", "Kıbrıscık", "Mengen", "Merkez", "Mudurnu", "Seben", "Yeniçağa"], "neighborhoods": ["Bolu OSB", "Gerede OSB"]},
    "Burdur": {"districts": ["Ağlasun", "Altınyayla", "Bucak", "Çavdır", "Çeltikçi", "Gölhisar", "Karamanlı", "Kemer", "Merkez", "Tefenni", "Yeşilova"], "neighborhoods": ["Bucak OSB"]},
    "Bursa": {"districts": ["Büyükorhan", "Gemlik", "Gürsu", "Harmancık", "İnegöl", "İznik", "Karacabey", "Keles", "Kestel", "Mudanya", "Mustafakemalpaşa", "Nilüfer", "Orhaneli", "Orhangazi", "Osmangazi", "Yenişehir", "Yıldırım"], "neighborhoods": ["Demirtaş", "NOSAB", "Hasanağa OSB", "Kayapa", "Beşevler"]},
    "Çanakkale": {"districts": ["Ayvacık", "Bayramiç", "Biga", "Bozcaada", "Çan", "Eceabat", "Ezine", "Gelibolu", "Gökçeada", "Lapseki", "Merkez", "Yenice"], "neighborhoods": ["Biga OSB", "Çanakkale OSB"]},
    "Çankırı": {"districts": ["Atkaracalar", "Bayramören", "Çerkeş", "Eldivan", "Ilgaz", "Kızılırmak", "Korgun", "Kurşunlu", "Merkez", "Orta", "Şabanözü", "Yapraklı"], "neighborhoods": ["Çerkeş OSB", "Şabanözü OSB"]},
    "Çorum": {"districts": ["Alaca", "Bayat", "Boğazkale", "Dodurga", "İskilip", "Kargı", "Laçin", "Mecitözü", "Merkez", "Oğuzlar", "Ortaköy", "Osmancık", "Sungurlu", "Uğurludağ"], "neighborhoods": ["Çorum OSB", "Sungurlu OSB"]},
    "Denizli": {"districts": ["Acıpayam", "Babadağ", "Baklan", "Bekilli", "Beyağaç", "Bozkurt", "Buldan", "Çal", "Çameli", "Çardak", "Çivril", "Güney", "Honaz", "Kale", "Merkezefendi", "Pamukkale", "Sarayköy", "Serinhisar", "Tavas"], "neighborhoods": ["Denizli OSB", "Honaz OSB"]},
    "Diyarbakır": {"districts": ["Bağlar", "Bismil", "Çermik", "Çınar", "Çüngüş", "Dicle", "Eğil", "Ergani", "Hani", "Hazro", "Kayapınar", "Kocaköy", "Kulp", "Lice", "Silvan", "Sur", "Yenişehir"], "neighborhoods": ["Diyarbakır OSB"]},
    "Düzce": {"districts": ["Akçakoca", "Cumayeri", "Çilimli", "Gölyaka", "Gümüşova", "Kaynaşlı", "Merkez", "Yığılca"], "neighborhoods": ["Düzce OSB", "Gümüşova OSB"]},
    "Edirne": {"districts": ["Enez", "Havsa", "İpsala", "Keşan", "Lalapaşa", "Meriç", "Merkez", "Süloğlu", "Uzunköprü"], "neighborhoods": ["Edirne OSB"]},
    "Elazığ": {"districts": ["Ağın", "Alacakaya", "Arıcak", "Baskil", "Karakoçan", "Keban", "Kovancılar", "Maden", "Merkez", "Palu", "Sivrice"], "neighborhoods": ["Elazığ OSB"]},
    "Erzincan": {"districts": ["Çayırlı", "İliç", "Kemah", "Kemaliye", "Merkez", "Otlukbeli", "Refahiye", "Tercan", "Üzümlü"], "neighborhoods": ["Erzincan OSB"]},
    "Erzurum": {"districts": ["Aşkale", "Aziziye", "Çat", "Hınıs", "Horasan", "İspir", "Karaçoban", "Karayazı", "Köprüköy", "Narman", "Oltu", "Olur", "Palandöken", "Pasinler", "Pazaryolu", "Şenkaya", "Tekman", "Tortum", "Uzundere", "Yakutiye"], "neighborhoods": ["Erzurum OSB"]},
    "Eskişehir": {"districts": ["Alpu", "Beylikova", "Çifteler", "Günyüzü", "Han", "İnönü", "Mahmudiye", "Mihalgazi", "Mihalıççık", "Odunpazarı", "Sarıcakaya", "Seyitgazi", "Sivrihisar", "Tepebaşı"], "neighborhoods": ["Eskişehir OSB", "Teksan", "Baksan"]},
    "Gaziantep": {"districts": ["Araban", "İslahiye", "Karkamış", "Nizip", "Nurdağı", "Oğuzeli", "Şahinbey", "Şehitkamil", "Yavuzeli"], "neighborhoods": ["Başpınar", "Gaziantep OSB", "KÜSGET", "Ünaldı"]},
    "Giresun": {"districts": ["Alucra", "Bulancak", "Çamoluk", "Çanakçı", "Dereli", "Doğankent", "Espiye", "Eynesil", "Görele", "Güce", "Keşap", "Merkez", "Piraziz", "Şebinkarahisar", "Tirebolu", "Yağlıdere"], "neighborhoods": ["Giresun OSB"]},
    "Gümüşhane": {"districts": ["Kelkit", "Köse", "Kürtün", "Merkez", "Şiran", "Torul"], "neighborhoods": []},
    "Hakkari": {"districts": ["Çukurca", "Derecik", "Merkez", "Şemdinli", "Yüksekova"], "neighborhoods": []},
    "Hatay": {"districts": ["Altınözü", "Antakya", "Arsuz", "Belen", "Defne", "Dörtyol", "Erzin", "Hassa", "İskenderun", "Kırıkhan", "Kumlu", "Payas", "Reyhanlı", "Samandağ", "Yayladağı"], "neighborhoods": ["Payas OSB", "İskenderun OSB", "Antakya OSB"]},
    "Iğdır": {"districts": ["Aralık", "Karakoyunlu", "Merkez", "Tuzluca"], "neighborhoods": ["Iğdır OSB"]},
    "Isparta": {"districts": ["Aksu", "Atabey", "Eğirdir", "Gelendost", "Gönen", "Keçiborlu", "Merkez", "Senirkent", "Sütçüler", "Şarkikaraağaç", "Uluborlu", "Yalvaç", "Yenişarbademli"], "neighborhoods": ["Isparta OSB"]},
    "İstanbul": {"districts": ["Adalar", "Arnavutköy", "Ataşehir", "Avcılar", "Bağcılar", "Bahçelievler", "Bakırköy", "Başakşehir", "Bayrampaşa", "Beşiktaş", "Beykoz", "Beylikdüzü", "Beyoğlu", "Büyükçekmece", "Çatalca", "Çekmeköy", "Esenler", "Esenyurt", "Eyüpsultan", "Fatih", "Gaziosmanpaşa", "Güngören", "Kadıköy", "Kağıthane", "Kartal", "Küçükçekmece", "Maltepe", "Pendik", "Sancaktepe", "Sarıyer", "Silivri", "Sultanbeyli", "Sultangazi", "Şile", "Şişli", "Tuzla", "Ümraniye", "Üsküdar", "Zeytinburnu"], "neighborhoods": ["İkitelli", "Dudullu", "Merter", "Maslak", "Levent", "Hadımköy", "Tuzla OSB", "İMES", "Perpa"]},
    "İzmir": {"districts": ["Aliağa", "Balçova", "Bayındır", "Bayraklı", "Bergama", "Bornova", "Buca", "Çeşme", "Çiğli", "Dikili", "Foça", "Gaziemir", "Güzelbahçe", "Karabağlar", "Karaburun", "Karşıyaka", "Kemalpaşa", "Kınık", "Kiraz", "Konak", "Menderes", "Menemen", "Narlıdere", "Ödemiş", "Seferihisar", "Selçuk", "Tire", "Torbalı", "Urla"], "neighborhoods": ["Işıkkent", "Pınarbaşı", "Atatürk OSB", "Kemalpaşa OSB"]},
    "Kahramanmaraş": {"districts": ["Afşin", "Andırın", "Çağlayancerit", "Dulkadiroğlu", "Ekinözü", "Elbistan", "Göksun", "Nurhak", "Onikişubat", "Pazarcık", "Türkoğlu"], "neighborhoods": ["Kahramanmaraş OSB", "Türkoğlu OSB"]},
    "Karabük": {"districts": ["Eflani", "Eskipazar", "Merkez", "Ovacık", "Safranbolu", "Yenice"], "neighborhoods": ["Karabük OSB"]},
    "Karaman": {"districts": ["Ayrancı", "Başyayla", "Ermenek", "Kazımkarabekir", "Merkez", "Sarıveliler"], "neighborhoods": ["Karaman OSB"]},
    "Kars": {"districts": ["Akyaka", "Arpaçay", "Digor", "Kağızman", "Merkez", "Sarıkamış", "Selim", "Susuz"], "neighborhoods": []},
    "Kastamonu": {"districts": ["Abana", "Ağlı", "Araç", "Azdavay", "Bozkurt", "Cide", "Çatalzeytin", "Daday", "Devrekani", "Doğanyurt", "Hanönü", "İhsangazi", "İnebolu", "Küre", "Merkez", "Pınarbaşı", "Seydiler", "Şenpazar", "Taşköprü", "Tosya"], "neighborhoods": ["Kastamonu OSB", "Tosya OSB"]},
    "Kayseri": {"districts": ["Akkışla", "Bünyan", "Develi", "Felahiye", "Hacılar", "İncesu", "Kocasinan", "Melikgazi", "Özvatan", "Pınarbaşı", "Sarıoğlan", "Sarız", "Talas", "Tomarza", "Yahyalı", "Yeşilhisar"], "neighborhoods": ["Kayseri OSB", "Mimarsinan OSB", "İncesu OSB"]},
    "Kırıkkale": {"districts": ["Bahşılı", "Balışeyh", "Çelebi", "Delice", "Karakeçili", "Keskin", "Merkez", "Sulakyurt", "Yahşihan"], "neighborhoods": ["Kırıkkale OSB"]},
    "Kırklareli": {"districts": ["Babaeski", "Demirköy", "Kofçaz", "Lüleburgaz", "Merkez", "Pehlivanköy", "Pınarhisar", "Vize"], "neighborhoods": ["Lüleburgaz OSB", "Kırklareli OSB"]},
    "Kırşehir": {"districts": ["Akçakent", "Akpınar", "Boztepe", "Çiçekdağı", "Kaman", "Merkez", "Mucur"], "neighborhoods": ["Kırşehir OSB"]},
    "Kilis": {"districts": ["Elbeyli", "Merkez", "Musabeyli", "Polateli"], "neighborhoods": ["Kilis OSB"]},
    "Kocaeli": {"districts": ["Başiskele", "Çayırova", "Darıca", "Derince", "Dilovası", "Gebze", "Gölcük", "İzmit", "Kandıra", "Karamürsel", "Kartepe", "Körfez"], "neighborhoods": ["Gebze OSB", "Dilovası OSB", "TOSB", "Arslanbey", "Alikahya"]},
    "Konya": {"districts": ["Ahırlı", "Akören", "Akşehir", "Altınekin", "Beyşehir", "Bozkır", "Cihanbeyli", "Çeltik", "Çumra", "Derbent", "Derebucak", "Doğanhisar", "Emirgazi", "Ereğli", "Güneysınır", "Hadim", "Halkapınar", "Hüyük", "Ilgın", "Kadınhanı", "Karapınar", "Karatay", "Kulu", "Meram", "Sarayönü", "Selçuklu", "Seydişehir", "Taşkent", "Tuzlukçu", "Yalıhüyük", "Yunak"], "neighborhoods": ["Büsan", "Konsan", "Horozluhan", "Konya OSB"]},
    "Kütahya": {"districts": ["Altıntaş", "Aslanapa", "Çavdarhisar", "Domaniç", "Dumlupınar", "Emet", "Gediz", "Hisarcık", "Merkez", "Pazarlar", "Simav", "Şaphane", "Tavşanlı"], "neighborhoods": ["Kütahya OSB", "Tavşanlı OSB"]},
    "Malatya": {"districts": ["Akçadağ", "Arapgir", "Arguvan", "Battalgazi", "Darende", "Doğanşehir", "Doğanyol", "Hekimhan", "Kale", "Kuluncak", "Pütürge", "Yazıhan", "Yeşilyurt"], "neighborhoods": ["Malatya OSB"]},
    "Manisa": {"districts": ["Ahmetli", "Akhisar", "Alaşehir", "Demirci", "Gölmarmara", "Gördes", "Kırkağaç", "Köprübaşı", "Kula", "Salihli", "Sarıgöl", "Saruhanlı", "Selendi", "Soma", "Şehzadeler", "Turgutlu", "Yunusemre"], "neighborhoods": ["Manisa OSB", "Turgutlu OSB", "Akhisar OSB"]},
    "Mardin": {"districts": ["Artuklu", "Dargeçit", "Derik", "Kızıltepe", "Mazıdağı", "Midyat", "Nusaybin", "Ömerli", "Savur", "Yeşilli"], "neighborhoods": ["Mardin OSB", "Kızıltepe OSB"]},
    "Mersin": {"districts": ["Akdeniz", "Anamur", "Aydıncık", "Bozyazı", "Çamlıyayla", "Erdemli", "Gülnar", "Mezitli", "Mut", "Silifke", "Tarsus", "Toroslar", "Yenişehir"], "neighborhoods": ["Kazanlı", "Karaduvar", "Serbest Bölge", "Tarsus OSB", "Mersin OSB"]},
    "Muğla": {"districts": ["Bodrum", "Dalaman", "Datça", "Fethiye", "Kavaklıdere", "Köyceğiz", "Marmaris", "Menteşe", "Milas", "Ortaca", "Seydikemer", "Ula", "Yatağan"], "neighborhoods": ["Milas OSB"]},
    "Muş": {"districts": ["Bulanık", "Hasköy", "Korkut", "Malazgirt", "Merkez", "Varto"], "neighborhoods": ["Muş OSB"]},
    "Nevşehir": {"districts": ["Acıgöl", "Avanos", "Derinkuyu", "Gülşehir", "Hacıbektaş", "Kozaklı", "Merkez", "Ürgüp"], "neighborhoods": ["Nevşehir OSB"]},
    "Niğde": {"districts": ["Altunhisar", "Bor", "Çamardı", "Çiftlik", "Merkez", "Ulukışla"], "neighborhoods": ["Niğde OSB", "Bor OSB"]},
    "Ordu": {"districts": ["Akkuş", "Altınordu", "Aybastı", "Çamaş", "Çatalpınar", "Çaybaşı", "Fatsa", "Gölköy", "Gülyalı", "Gürgentepe", "İkizce", "Kabadüz", "Kabataş", "Korgan", "Kumru", "Mesudiye", "Perşembe", "Ulubey", "Ünye"], "neighborhoods": ["Fatsa OSB", "Ordu OSB"]},
    "Osmaniye": {"districts": ["Bahçe", "Düziçi", "Hasanbeyli", "Kadirli", "Merkez", "Sumbas", "Toprakkale"], "neighborhoods": ["Osmaniye OSB", "Kadirli OSB"]},
    "Rize": {"districts": ["Ardeşen", "Çamlıhemşin", "Çayeli", "Derepazarı", "Fındıklı", "Güneysu", "Hemşin", "İkizdere", "İyidere", "Kalkandere", "Merkez", "Pazar"], "neighborhoods": ["Rize OSB"]},
    "Sakarya": {"districts": ["Adapazarı", "Akyazı", "Arifiye", "Erenler", "Ferizli", "Geyve", "Hendek", "Karapürçek", "Karasu", "Kaynarca", "Kocaali", "Pamukova", "Sapanca", "Serdivan", "Söğütlü", "Taraklı"], "neighborhoods": ["Sakarya OSB", "Arifiye", "Hanlı", "Erenler"]},
    "Samsun": {"districts": ["Alaçam", "Asarcık", "Atakum", "Ayvacık", "Bafra", "Canik", "Çarşamba", "Havza", "İlkadım", "Kavak", "Ladik", "Ondokuzmayıs", "Salıpazarı", "Tekkeköy", "Terme", "Vezirköprü", "Yakakent"], "neighborhoods": ["Tekkeköy", "Samsun OSB", "Kavak OSB"]},
    "Siirt": {"districts": ["Baykan", "Eruh", "Kurtalan", "Merkez", "Pervari", "Şirvan", "Tillo"], "neighborhoods": ["Siirt OSB"]},
    "Sinop": {"districts": ["Ayancık", "Boyabat", "Dikmen", "Durağan", "Erfelek", "Gerze", "Merkez", "Saraydüzü", "Türkeli"], "neighborhoods": ["Sinop OSB"]},
    "Sivas": {"districts": ["Akıncılar", "Altınyayla", "Divriği", "Doğanşar", "Gemerek", "Gölova", "Gürün", "Hafik", "İmranlı", "Kangal", "Koyulhisar", "Merkez", "Suşehri", "Şarkışla", "Ulaş", "Yıldızeli", "Zara"], "neighborhoods": ["Sivas OSB"]},
    "Şanlıurfa": {"districts": ["Akçakale", "Birecik", "Bozova", "Ceylanpınar", "Eyyübiye", "Halfeti", "Haliliye", "Harran", "Hilvan", "Karaköprü", "Siverek", "Suruç", "Viranşehir"], "neighborhoods": ["Şanlıurfa OSB"]},
    "Şırnak": {"districts": ["Beytüşşebap", "Cizre", "Güçlükonak", "İdil", "Merkez", "Silopi", "Uludere"], "neighborhoods": ["Şırnak OSB", "Cizre OSB"]},
    "Tekirdağ": {"districts": ["Çerkezköy", "Çorlu", "Ergene", "Hayrabolu", "Kapaklı", "Malkara", "Marmaraereğlisi", "Muratlı", "Saray", "Süleymanpaşa", "Şarköy"], "neighborhoods": ["Çerkezköy OSB", "Çorlu OSB", "Ergene OSB", "Kapaklı OSB"]},
    "Tokat": {"districts": ["Almus", "Artova", "Başçiftlik", "Erbaa", "Merkez", "Niksar", "Pazar", "Reşadiye", "Sulusaray", "Turhal", "Yeşilyurt", "Zile"], "neighborhoods": ["Tokat OSB", "Erbaa OSB"]},
    "Trabzon": {"districts": ["Akçaabat", "Araklı", "Arsin", "Beşikdüzü", "Çarşıbaşı", "Çaykara", "Dernekpazarı", "Düzköy", "Hayrat", "Köprübaşı", "Maçka", "Of", "Ortahisar", "Sürmene", "Şalpazarı", "Tonya", "Vakfıkebir", "Yomra"], "neighborhoods": ["Arsin OSB", "Beşikdüzü OSB"]},
    "Tunceli": {"districts": ["Çemişgezek", "Hozat", "Mazgirt", "Merkez", "Nazımiye", "Ovacık", "Pertek", "Pülümür"], "neighborhoods": []},
    "Uşak": {"districts": ["Banaz", "Eşme", "Karahallı", "Merkez", "Sivaslı", "Ulubey"], "neighborhoods": ["Uşak OSB", "Karma OSB"]},
    "Van": {"districts": ["Bahçesaray", "Başkale", "Çaldıran", "Çatak", "Edremit", "Erciş", "Gevaş", "Gürpınar", "İpekyolu", "Muradiye", "Özalp", "Saray", "Tuşba"], "neighborhoods": ["Van OSB"]},
    "Yalova": {"districts": ["Altınova", "Armutlu", "Çınarcık", "Çiftlikköy", "Merkez", "Termal"], "neighborhoods": ["Yalova OSB"]},
    "Yozgat": {"districts": ["Akdağmadeni", "Aydıncık", "Boğazlıyan", "Çandır", "Çayıralan", "Çekerek", "Kadışehri", "Merkez", "Saraykent", "Sarıkaya", "Sorgun", "Şefaatli", "Yenifakılı", "Yerköy"], "neighborhoods": ["Yozgat OSB", "Sorgun OSB"]},
    "Zonguldak": {"districts": ["Alaplı", "Çaycuma", "Devrek", "Ereğli", "Gökçebey", "Kilimli", "Kozlu", "Merkez"], "neighborhoods": ["Çaycuma OSB", "Ereğli OSB"]},
}

TURKEY_CITIES = list(LOCATION_DICTIONARY.keys())

# v16: İlçe -> il listesi. Aynı ilçe adı birden çok ilde geçiyorsa
# il çıkarımı için tek başına kullanma; şehir/district yakın bağlamı iste.
DISTRICT_TO_CITIES = {}
for _city, _data in LOCATION_DICTIONARY.items():
    for _district in _data.get("districts", []):
        DISTRICT_TO_CITIES.setdefault(normalize_text(_district), []).append(_city)

def district_is_unique(district):
    return len(DISTRICT_TO_CITIES.get(normalize_text(district), [])) == 1

# İlçe isimlerinden bazıları günlük Türkçede de kullanılan kelimeler.
# Bunlar tek başına geçince ilçe sayılmaz; açık ilçe bağlamı gerekir.
# v10: Şehir adı tek başına yeterli değildir. Örn. "Bursa'da yıldırım düştü" -> Yıldırım ilçesi değildir.
AMBIGUOUS_DISTRICT_NAMES = {
    "yildirim",   # Bursa/Yıldırım ama "yıldırım düştü" olabilir
    "arac",       # Kastamonu/Araç ama "8 araç" olabilir
    "of",         # Trabzon/Of ama kısa/gürültülü
    "pazar",      # Pazar kelimesi genel kullanım
    "kale",       # Kale kelimesi genel kullanım
    "merkez",     # Her ilde var, tek başına lokasyon değeri düşük
    "han",        # Eskişehir/Han ama genel kelime
    "mut",        # Mersin/Mut ama kısa/riskli
    "sur",        # Diyarbakır/Sur ama genel/özel isim olabilir
    "genc",       # Bingöl/Genç ama genel kelime
    "maden",      # Elazığ/Maden ama sektör kelimesi
    "kemer",      # Antalya/Kemer ama genel kelime
    "cay",        # Afyon/Çay ama genel kelime
    "aksu",       # Aksu çok yerde mahalle/akarsu adı olabilir
    "saray",      # Saray genel kelime/özel isim
    "havza",      # Havza genel kelime
    "termal",     # Termal genel kelime
    "fatih",      # kişi adı olarak çok geçer
    "gonen",      # fiil/özel isim gürültüsü olabilir
    "evciler",    # genel ad gibi geçebilir
    "koprubasi",  # yer/mahalle/genel ifade
    "pinarbasi",  # yer/mahalle/genel ifade
    "ulubey",     # kişi/soyad/marka gibi geçebilir
    "yenisehir",  # mahalle/yer adı olarak çok geçer
    "can",         # Çanakkale/Çan ama "can kaybı" gibi günlük kelime
    "imamoglu",    # Adana/İmamoğlu ama siyasi kişi/haber olarak çok geçer
    "orta",        # Çankırı/Orta ama "orta boy", "orta ölçekli" gibi geçer
    "olur",        # Erzurum/Olur ama fiil olarak çok geçer
    "cinar",       # Diyarbakır/Çınar ama haber sitesi/soyad gibi geçer
    "kazan",       # Ankara/Kazan ama fiil/isim olarak geçer
}

# İlçe gibi görünen ama olay anlamı taşıyan bağlamlar.
DISTRICT_EVENT_CONTEXT = {
    "yildirim": [
        "yildirim dustu", "yildirim dusmesi", "yildirim carpmasi", "yildirim isabet",
        "yildirim kaynakli", "yildirim nedeniyle", "yildirim sonucu", "yildirimdan"
    ],
    "arac": [
        "arac zarar", "arac yandi", "arac yangini", "arac kullan", "arac sahibi",
        "araclar", "aracin", "araca", "aracta"
    ],
    "pazar": ["pazar gunu", "pazar yeri", "pazar payi", "pazar arastirmasi"],
    "maden": ["maden ocagi", "maden sahasi", "maden sektoru"],
    "kale": ["kale gibi", "kale diregi"],
    "merkez": ["merkez bankasi", "merkez uretim", "merkez ofis"],
    "cay": ["cay bahcesi", "cay uretimi", "cay fabrikasi"],
    "can": ["can kaybi", "can kaybı", "can guvenligi", "can güvenliği", "can pazari", "can pazarı"],
    "imamoglu": ["imamoglu fondas", "imamoglu", "ekrem imamoglu"],
    "orta": ["orta boy", "orta olcekli", "orta ölçekli", "orta vadeli", "orta gelir"],
    "olur": ["olur mu", "olur olmaz", "olur dedi", "olur ise", "olur gibi"],
    "cinar": ["kocaeli cinar", "cinar haber", "cinar gazetesi"],
    "kazan": ["kazan dairesi", "kazan patlamasi", "kazan patlaması", "kazan yakiti"],
}


def has_event_context_for_ambiguous_district(original_text, district):
    norm = normalize_text(original_text)
    d_norm = normalize_text(district)
    for phrase in DISTRICT_EVENT_CONTEXT.get(d_norm, []):
        if phrase in norm:
            return True
    return False


def has_explicit_district_context(original_text, district):
    """
    Riskli ilçelerde kabul edilen güçlü bağlamlar:
    - Bandırma'da / Bandırma’da / Bandırma ilçesinde
    - Bursa Yıldırım / Yıldırım Bursa gibi şehir-ilçe yakınlığı
    - Yıldırım OSB / Araç Organize Sanayi gibi açık lokasyon bağlamı
    """
    text = str(original_text)
    d = re.escape(str(district))
    patterns = [
        rf"\b{d}\s*['’]?\s*(da|de|ta|te|dan|den|tan|ten)\b",
        rf"\b{d}\s+(ilçesi|ilcesi|ilçesinde|ilcesinde|ilçesine|ilcesine|ilçeden|ilceden|ilçe|ilce)\b",
        rf"\b{d}\s+(osb|organize sanayi|sanayi sitesi|mahallesi|caddesi|sokağı|sokagi)\b",
    ]
    return any(re.search(pat, text, flags=re.IGNORECASE) for pat in patterns)


def has_city_district_pair_context(original_text, city, district):
    """
    Şehir ve riskli ilçe isimleri aynı metinde ama uzaksa kabul etmiyoruz.
    Yakın bağlam örnekleri:
    - Bursa Yıldırım
    - Bursa'nın Yıldırım ilçesi
    - Yıldırım/Bursa
    """
    city_n = normalize_text(city)
    district_n = normalize_text(district)
    norm = normalize_text(original_text)

    patterns = [
        rf"\b{re.escape(city_n)}\s+(?:ili\s+|ilinde\s+|ilinin\s+|merkez\s+)?{re.escape(district_n)}\b",
        rf"\b{re.escape(city_n)}\s+.{0,30}\b{re.escape(district_n)}\s+(?:ilcesi|ilcesinde|ilce)\b",
        rf"\b{re.escape(district_n)}\s*/\s*{re.escape(city_n)}\b",
        rf"\b{re.escape(city_n)}\s*/\s*{re.escape(district_n)}\b",
    ]
    return any(re.search(pat, norm) for pat in patterns)


def district_match_is_safe(original_text, text_norm, city, district):
    district_norm = normalize_text(district)

    if re.search(rf"\b{re.escape(district_norm)}\b", text_norm) is None:
        return False

    if has_event_context_for_ambiguous_district(original_text, district):
        return False

    # v20: Artık ilçe adı sözlükte var diye otomatik kabul edilmiyor.
    # Maltepe'de / Burhaniye ilçesinde / Balıkesir-Burhaniye gibi açık lokasyon bağlamı aranıyor.
    explicit = has_district_explicit_context_any(original_text, district)
    pair = has_city_district_pair_context(original_text, city, district)

    if district_norm in AMBIGUOUS_DISTRICT_NAMES:
        return explicit or pair

    # Non-ambiguous ilçelerde de en az explicit bağlam veya şehir-ilçe yakınlığı şart.
    return explicit or pair


# ============================================================
# NOISE WORDS
# ============================================================

NEWS_SOURCE_WORDS = [
    "haber", "haberleri", "gazete", "gazetesi", "son dakika",
    "gündem", "ekonomi", "ekonomi haberleri", "video", "foto galeri",
    "galeri", "canlı", "son gelişme", "son gelişmeler",
    "hürriyet", "sabah", "milliyet", "cumhuriyet", "sözcü", "sozcu",
    "habertürk", "haberturk", "ntv", "cnn türk", "cnn turk",
    "trt haber", "yeni şafak", "yenisafak", "karar", "takvim",
    "akşam", "aksam", "ensonhaber", "haber7", "haber 7", "gzt",
    "ekonomim", "patronlar dünyası", "patronlar dunyasi",
    "emlak kulisi", "ege telgraf", "iha", "dha", "anka",
    "anadolu ajansı", "anadolu ajansi", "demirören haber ajansı",
    "demiroren haber ajansi", "ihlas haber ajansı", "ihlas haber ajansi",
    "bursahaber", "odatv", "superhaber", "medyaradar",
    "memleket", "bölge gündemi", "bolge gundemi", "manset aydin",
    "manşet aydın", "haber aktüel", "haber aktuel", "milli gazete",
    "korkusuz", "yeni akit", "akasyam", "sondakika", "12punto",
    "10haber", "t24", "birgün", "birgun", "evrensel", "avazturk",
    "gerçek haberci", "gercek haberci", "internethaber", "internet haber"
]

GENERAL_NOISE_WORDS = [
    "son", "dakika", "flaş", "flas", "şok", "sok", "korkutan",
    "büyük", "buyuk", "işte", "iste", "detay", "detaylar",
    "açıklandı", "aciklandi", "iddia", "iddiaları", "iddialari",
    "gelişme", "gelisme", "olay", "panik", "alev", "alevler",
    "yükseldi", "yukseldi", "kontrol", "altına", "altina",
    "alındı", "alindi", "söndürüldü", "sonduruldu", "çıktı",
    "cikti", "meydana", "gelen", "yaşandı", "yasandi", "edildi",
    "oldu", "olduğu", "oldugu", "başladı", "basladi", "dev",
    "ünlü", "unlu", "yıllık", "yillik", "köklü", "koklu"
]

BAD_CANDIDATE_PHRASES = [
    "yillik dev", "dev gida", "dev tekstil", "iflas etti",
    "konkordato istedi", "konkordato talep etti", "iflasin esiginde",
    "sene sonuna kadar", "yuzlerce magaza", "sektorde zincirleme",
    "te is yeri", "de is yeri", "nin tekstil", "gida turkiye",
    "tekstil turkiye", "isletme turkiye", "gazetesi", "gazete",
    "haberleri", "ekonomi haberleri", "de tekstil", "ler mersin",
    "devini daha", "iflasa surukledi", "daha iflasa",
    "turkiye turkiye", "arama sonuc", "google maps"
]

COMMERCIAL_ENTITY_PATTERN = (
    r"(a\.?\s*ş\.?|ltd|limited|anonim|holding|sanayi|ticaret|"
    r"tekstil|gıda|gida|lojistik|enerji|inşaat|insaat|otomotiv|"
    r"akaryakıt|akaryakit|fabrika|depo|tesis|kuruluş|kurulusu|"
    r"kimya|metal|makina|makine|plastik|ambalaj|elektrik|elektronik|"
    r"madencilik|maden|tarım|tarim|hayvancılık|hayvancilik|ilaç|ilac|"
    r"sağlık|saglik|petrol|doğalgaz|dogalgaz|mobilya|giyim|ayakkabı|ayakkabi|"
    r"kağıt|kagit|cam|seramik|çimento|cimento|demir|çelik|celik)"
)

# ============================================================
# ENTITY QUALITY FILTERS - v7
# ============================================================

DISTRICT_BLACKLIST = [
    "yangin", "yangini", "yanginin", "yanginda", "dava", "davasi", "davasinda",
    "kritik", "durusma", "haber", "haberleri", "ajans", "ajansi", "com",
    "son", "dakika", "gundem", "ekonomi", "dunya", "turkiye", "gazete",
    "gazetesi", "mobilya", "fabrika", "fabrikasi", "depo", "tesisi",
    "alindi", "bandirma", "yangini davasin", "ajansi maltepe", "dakika bandirma"
]

ENTITY_EXTRACTION_BLACKLIST = [
    "davasinda", "dava", "kritik", "durusma", "haber", "haberleri", "gazete",
    "gazetesi", "son dakika", "com", "www", "http", "https", "ajansi",
    "yangini davasinda", "kritik durusma", "hangi sirket", "devini",
    "iflasa surukledi", "son verildi", "kisinin isine", "kişinin işine"
]

VALID_COMPANY_SECTORS = [
    "Tekstil", "İnşaat", "Gıda", "Otomotiv", "Sanayi", "Ticaret", "Lojistik",
    "Enerji", "Holding", "Akaryakıt", "Mobilya", "Turizm", "Kimya", "Metal",
    "Makina", "Makine", "Plastik", "Ambalaj", "Elektrik", "Elektronik",
    "Madencilik", "Maden", "Tarım", "Hayvancılık", "İlaç", "Sağlık",
    "Petrol", "Doğalgaz", "Giyim", "Ayakkabı", "Kağıt", "Cam", "Seramik",
    "Çimento", "Demir", "Çelik", "Tavukçuluk", "Tavukculuk"
]

FACILITY_SUFFIXES = [
    "fabrikası", "fabrikasi", "deposu", "tesisi", "atölyesi", "atolyesi",
    "akaryakıt istasyonu", "akaryakit istasyonu", "şantiyesi", "santiyesi",
    "üretim tesisi", "uretim tesisi", "lojistik deposu", "dağıtım merkezi",
    "dagitim merkezi", "dolum tesisi"
]

# ============================================================
# ARTICLE / SIDEBAR CLEANING - v17
# ============================================================

# Gerçek haber sayfalarında ham metne "ilgili haberler", "en çok okunanlar",
# sidebar, footer ve galeri navigasyonu karışabiliyor. raw_text Excel'de kalır;
# entity extraction için ayrıca temiz bir kaynak metin üretilir.
SIDEBAR_CUT_MARKERS = [
    "ilgili haberler", "benzer haberler", "sıradaki haber", "siradaki haber",
    "en çok okunan", "en cok okunan", "çok okunan", "cok okunan",
    "son dakika haberleri", "son dakika", "günün manşetleri", "gunun mansetleri",
    "öne çıkan haberler", "one cikan haberler", "editörün seçtikleri", "editorun sectikleri",
    "tavsiye edilen içerikler", "tavsiye edilen icerikler", "diğer haberler", "diger haberler",
    "bunlar da ilginizi çekebilir", "bunlar da ilginizi cekebilir",
    "sıradaki galeri", "siradaki galeri", "foto galeri", "video galeri",
    "yorumlar", "etiketler", "paylaş", "paylas", "abone ol", "çerez", "cerez",
    "gözaltına alınan bazı isimler", "gozaltina alinan bazi isimler",
    "bazı isimler şu şekilde", "bazi isimler su sekilde", "şu şekilde:", "su sekilde:",
    "gözaltı listesinde", "gozalti listesinde",
]

ENTITY_RELEVANT_KEYWORDS = [
    "iflas", "konkordato", "tasfiye", "yangın", "yangin", "fabrika", "depo", "tesis",
    "şirket", "sirket", "firma", "limited", "ltd", "a.ş", "aş", "anonim",
    "sanayi", "ticaret", "tavukçuluk", "tavukculuk", "gıda", "gida", "lojistik",
    "balıkesir", "balikesir", "burhaniye", "çanakkale", "canakkale", "adana",
    "masak", "kara para", "dolandırıcılık", "dolandiricilik", "soruşturma", "sorusturma",
]

ENTITY_SENTENCE_KEEP_FIRST_N = 12
ENTITY_TEXT_MAX_SENTENCES = 80



def cut_sidebar_tail(text):
    """
    v17: Ham haber metninin sonunda gelen sidebar/ilgili haber/footer bloklarını keser.
    Çok erken geçen marker'ları kesmez; haberin ana gövdesinin bozulmasını engeller.
    """
    text = clean_spaces(text)
    if not text:
        return ""

    lower = text.lower()
    cut_positions = []
    min_cut = max(350, int(len(text) * 0.18))

    for marker in SIDEBAR_CUT_MARKERS:
        idx = lower.find(marker.lower())
        if idx >= min_cut:
            cut_positions.append(idx)

    if cut_positions:
        text = text[:min(cut_positions)]

    return clean_spaces(text)


def split_entity_sentences(text):
    text = clean_spaces(text)
    if not text:
        return []

    # v25: Ltd. Şti. / A.Ş. gibi legal kısaltmalar cümle bölünmesinde parçalanmasın.
    # Önce noktaları geçici tokenlara çevir, split sonrası geri al.
    protected = text
    replacements = {
        "Ltd. Şti.": "Ltd§ Şti§", "LTD. ŞTİ.": "LTD§ ŞTİ§",
        "Ltd.Şti.": "Ltd§Şti§", "LTD.ŞTİ.": "LTD§ŞTİ§",
        "A.Ş.": "A§Ş§", "A. Ş.": "A§ Ş§", "a.ş.": "a§ş§",
        "Şti.": "Şti§", "ŞTİ.": "ŞTİ§", "Ltd.": "Ltd§", "LTD.": "LTD§",
    }
    for k, v in replacements.items():
        protected = protected.replace(k, v)

    parts = re.split(r"(?<=[.!?])\s+|\s{2,}|\n+", protected)

    restored = []
    for part in parts:
        part = part.replace("§", ".")
        part = clean_spaces(part)
        if len(part) >= 25:
            restored.append(part)
    return restored


def significant_tokens(text):
    norm = normalize_text(text)
    stop = set([
        "haber", "son", "dakika", "icin", "olan", "oldu", "ile", "bir", "çok", "cok",
        "daha", "sonra", "gibi", "yeni", "gundem", "gündem", "ekonomi", "turkiye", "türkiye",
        "com", "www", "video", "galeri"
    ])
    return {t for t in norm.split() if len(t) >= 4 and t not in stop}


def sentence_relevance_score(sentence, title_tokens, summary_tokens):
    norm = normalize_text(sentence)
    toks = set(norm.split())
    score = 0

    if title_tokens:
        score += len(toks & title_tokens) * 8
    if summary_tokens:
        score += len(toks & summary_tokens) * 5

    for kw in ENTITY_RELEVANT_KEYWORDS:
        if normalize_text(kw) in norm:
            score += 12

    # Legal suffix / gerçek şirket formatı güçlü sinyal.
    if re.search(r"\b(a\.?\s*ş\.?|ltd\.?\s*şti\.?|limited\s+şirketi|anonim\s+şirketi)\b", sentence, flags=re.IGNORECASE):
        score += 35

    # İl / ilçe / OSB gibi lokasyon sinyali.
    if any(re.search(rf"\b{re.escape(normalize_text(city))}\b", norm) for city in TURKEY_CITIES):
        score += 10
    if any(x in norm for x in [" ilcesi", " ilcesinde", " mahallesi", " osb", " organize sanayi"]):
        score += 12

    # Sidebar cümleleri genelde kısa başlık kümeleri gibi olur.
    if any(normalize_text(m) in norm for m in SIDEBAR_CUT_MARKERS):
        score -= 35
    if has_bad_source_word(sentence):
        score -= 10

    return score


def build_entity_source_text(title, summary, raw_text):
    """
    v17: Entity extraction için kaynak metin üretir.
    raw_text tam olarak korunur; ancak entity çıkarımı sidebar/footer/ilgili haberlerden arındırılmış metinle yapılır.
    İlk 1200 karakter gibi sert kesme yoktur. Skor bazlı cümle seçimi vardır.
    """
    title = clean_spaces(title)
    summary = clean_spaces(summary)
    raw_text = cut_sidebar_tail(raw_text)

    title_tokens = significant_tokens(title)
    summary_tokens = significant_tokens(summary)
    sentences = split_entity_sentences(raw_text)

    kept = []
    seen = set()
    for i, sent in enumerate(sentences):
        norm = normalize_text(sent)
        if not norm or norm in seen:
            continue
        seen.add(norm)

        score = sentence_relevance_score(sent, title_tokens, summary_tokens)
        if i < ENTITY_SENTENCE_KEEP_FIRST_N:
            score += 18

        if score >= 10:
            kept.append((score, i, sent))

    # Metin uzun olsa bile entity için sadece haberle alakalı cümleler kullanılır.
    # Gövdenin sonundaki ilgili haber/sidebar başlıkları böylece ilçe/firma üretmez.
    kept.sort(key=lambda x: (x[1]))
    selected = [x[2] for x in kept[:ENTITY_TEXT_MAX_SENTENCES]]

    entity_text = clean_spaces(" ".join([title, summary] + selected))
    if len(entity_text) < len(clean_spaces(f"{title} {summary}")) + 50:
        # Çok agresif temizlik olduysa en azından kesilmiş raw_text'in başından makul bir parça ekle.
        entity_text = clean_spaces(f"{title} {summary} {' '.join(sentences[:ENTITY_SENTENCE_KEEP_FIRST_N])}")

    return entity_text

# ============================================================
# NEWS FUNCTIONS
# ============================================================

def clean_html(html):
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    return clean_spaces(soup.get_text(" "))



def extract_article_text_from_html(html, url=""):
    """
    v18 hızlı ve daha temiz içerik çıkarımı.
    Önce JSON-LD/meta/article/main/p etiketleriyle hızlı metin üretir.
    Trafılatūra sadece hızlı yöntem yetersiz kalırsa çalışır; bu fallback süresini ciddi azaltır.
    """
    html = html or ""

    try:
        soup = BeautifulSoup(html, "html.parser")
        for tag in soup(["script", "style", "noscript", "svg", "iframe", "form", "nav", "footer", "aside"]):
            tag.decompose()

        selector_texts = []
        seen = set()

        # JSON-LD articleBody çoğu haberde en temiz kaynaktır.
        for script in soup.find_all("script", {"type": "application/ld+json"}):
            raw = script.string or script.get_text(" ")
            for key in ["articleBody", "description", "headline"]:
                m = re.search(rf'"{key}"\s*:\s*"(.*?)"', raw, flags=re.DOTALL)
                if m:
                    txt = clean_spaces(m.group(1).encode("utf-8").decode("unicode_escape", errors="ignore"))
                    if len(txt) >= 80:
                        selector_texts.append(txt)

        for meta_sel in [
            {"property": "og:description"}, {"name": "description"}, {"name": "twitter:description"}
        ]:
            el = soup.find("meta", meta_sel)
            if el and el.get("content"):
                selector_texts.append(clean_spaces(el.get("content")))

        selectors = [
            "article", "main", "[itemprop='articleBody']",
            "div[class*='article']", "div[class*='content']", "div[class*='detail']", "div[class*='haber']",
            "div[class*='gallery']", "div[class*='galeri']", "figcaption",
            "h1", "h2", "p"
        ]

        for selector in selectors:
            for el in soup.select(selector):
                # Link yoğun bloklar genelde sidebar/related news'tir.
                txt = clean_spaces(el.get_text(" "))
                norm = normalize_text(txt)
                if len(txt) < 25 or norm in seen:
                    continue
                link_count = len(el.find_all("a")) if hasattr(el, "find_all") else 0
                p_count = len(el.find_all("p")) if hasattr(el, "find_all") else 0
                if link_count >= 5 and p_count <= 2:
                    continue
                if any(noise in norm for noise in ["cookie", "cerez", "giris yap", "abone ol", "reklam", "en cok okunan"]):
                    continue
                seen.add(norm)
                selector_texts.append(txt)

        selector_text = clean_spaces(" ".join(selector_texts))
        selector_text = cut_sidebar_tail(selector_text)
        if len(selector_text) >= MIN_ARTICLE_TEXT_LENGTH:
            return selector_text
    except Exception:
        pass

    # Son çare: trafilatura. Ağ değil CPU yavaşlığı yaratabildiği için sona alındı.
    if trafilatura is not None and len(html) < 1_500_000:
        try:
            extracted = trafilatura.extract(
                html,
                include_comments=False,
                include_tables=False,
                favor_precision=True,
                url=url
            )
            extracted = cut_sidebar_tail(clean_spaces(extracted or ""))
            if len(extracted) >= MIN_ARTICLE_TEXT_LENGTH:
                return extracted
        except Exception:
            pass

    text = cut_sidebar_tail(clean_html(html))
    return text if len(text) >= MIN_ARTICLE_TEXT_LENGTH else ""

def is_google_news_url(url):
    return "news.google.com" in normalize_text(url or "")


def fetch_article(url, force=False):
    """
    Requests tabanlı hızlı içerik denemesi.
    Google News RSS linklerinde çoğu zaman gerçek siteye tam geçemediği için
    v12'de bunun üstüne browser fallback de var.
    """
    if not force and not ENABLE_ARTICLE_FETCH:
        return "", ""

    if not url:
        return "", ""

    try:
        r = requests.get(url, headers=HEADERS, timeout=ARTICLE_FETCH_TIMEOUT, allow_redirects=True)
        if r.status_code != 200:
            return "", ""

        resolved_url = r.url or url
        article_text = extract_article_text_from_html(r.text, resolved_url)

        # Google News wrapper içinde kaldıysa browser fallback'e bırak.
        if is_google_news_url(resolved_url) and not article_text:
            return "", resolved_url

        return article_text, resolved_url

    except Exception:
        return "", ""


async def fetch_article_with_browser(page, url):
    """
    v21: v20'deki hızlı DOM-evaluate yöntemi bazı haberlerde boş/eksik metin döndürüp
    browser_failed üretiyordu. v17'de daha iyi çalışan yönteme geri dönüldü:
    - Google News linkini browser ile aç
    - commit sonrası kısa bekle
    - önce tam HTML'i alıp mevcut extract_article_text_from_html ile çöz
    - olmazsa DOM text fallback dene
    Böylece v17'de içerik çekebilen haberler yeniden çekilir.
    """
    if not url:
        return "", "", "browser_no_url"

    try:
        await page.goto(url, wait_until="commit", timeout=BROWSER_ARTICLE_GOTO_TIMEOUT_MS)
        await page.wait_for_timeout(BROWSER_ARTICLE_WAIT_MS)

        resolved_url = page.url

        # 1) v17 ile uyumlu ana yol: HTML'i al, mevcut HTML extractor'a ver.
        try:
            html = await page.content()
            article_text = extract_article_text_from_html(html, resolved_url)
            if article_text and len(article_text) >= MIN_ARTICLE_TEXT_LENGTH:
                return article_text, resolved_url, "browser_full_article"
        except Exception:
            article_text = ""

        # 2) Bazı sitelerde content() çok zayıf kalırsa DOM görünür metin fallback.
        try:
            visible_text = await page.evaluate(r"""
            () => {
                const kill = ['script','style','noscript','svg','iframe','form'];
                for (const sel of kill) document.querySelectorAll(sel).forEach(e => e.remove());
                const selectors = ['article','main','[itemprop="articleBody"]','div[class*="article"]','div[class*="content"]','div[class*="detail"]','div[class*="haber"]','div[class*="gallery"]','div[class*="galeri"]','figcaption','h1','h2','p'];
                const out = [];
                const seen = new Set();
                for (const sel of selectors) {
                    for (const el of document.querySelectorAll(sel)) {
                        const links = el.querySelectorAll ? el.querySelectorAll('a').length : 0;
                        const ps = el.querySelectorAll ? el.querySelectorAll('p').length : 0;
                        if (links >= 8 && ps <= 2) continue;
                        const t = (el.innerText || el.textContent || '').trim().replace(/\s+/g, ' ');
                        const key = t.toLowerCase();
                        if (t.length < 25 || seen.has(key)) continue;
                        if (/cookie|çerez|cerez|abone ol|giriş yap|giris yap|reklam/i.test(t)) continue;
                        seen.add(key);
                        out.push(t);
                    }
                }
                return out.join(' ');
            }
            """)
            visible_text = cut_sidebar_tail(clean_spaces(visible_text))
            if visible_text and len(visible_text) >= MIN_ARTICLE_TEXT_LENGTH:
                return visible_text, resolved_url, "browser_full_article_dom"
        except Exception:
            pass

        return "", resolved_url, "browser_failed"

    except Exception:
        return "", "", "browser_error"




def fallback_info_score(row):
    """
    v15: Bir haberin kontrol açısından ne kadar dolu olduğunu ölçer.
    Düşük skor = browser içerik fallback için daha iyi aday.
    """
    score = 0
    if str(row.get("city_candidates", "")).strip():
        score += 25
    if str(row.get("district_candidates", "")).strip():
        score += 20
    if str(row.get("company_candidates", "")).strip():
        score += 35
    if str(row.get("facility_candidates", "")).strip():
        score += 10
    if str(row.get("neighborhood_candidates", "")).strip():
        score += 5
    if len(str(row.get("raw_text", ""))) >= MIN_ARTICLE_TEXT_LENGTH:
        score += 5
    return score


def fallback_priority_score(row):
    """
    v15: Eksik kalan haberler arasında hangisine önce gidileceğini belirler.
    İflas/yangın gibi değerli olaylar ve çok eksik entity'ler öne alınır.
    """
    category = normalize_text(row.get("category", ""))
    title = normalize_text(row.get("title", ""))
    summary = normalize_text(row.get("summary", ""))
    text = f"{category} {title} {summary}"

    priority = 0
    if any(x in text for x in ["iflas", "konkordato", "mali sorun", "tasfiye"]):
        priority += 60
    if any(x in text for x in ["yangin", "fabrika", "depo", "tesis"]):
        priority += 45
    if any(x in text for x in ["masak", "kara para", "para aklama", "bahis", "dolandiricilik"]):
        priority += 40

    info = fallback_info_score(row)
    priority += max(0, 100 - info)

    if len(str(row.get("summary", ""))) < MIN_SUMMARY_LENGTH_FOR_NO_FETCH:
        priority += 20
    if len(str(row.get("raw_text", ""))) < MIN_ARTICLE_TEXT_LENGTH:
        priority += 20

    return priority


def needs_browser_fallback(row):
    """
    v15: Browser fallback sadece eksik bilgi kalan haberlerde çalışır.
    Amaç: 50 milyon yumurta / Gültav gibi değerli ama summary'de eksik kalan haberleri kurtarmak.
    """
    status = str(row.get("article_fetch_status", "summary_only"))
    raw_len = len(str(row.get("raw_text", "")))
    summary_len = len(str(row.get("summary", "")))

    already_full = status in ["requests_full_article", "browser_full_article"] and raw_len >= MIN_ARTICLE_TEXT_LENGTH
    if already_full:
        return False

    city = str(row.get("city_candidates", "")).strip()
    district = str(row.get("district_candidates", "")).strip()
    company = str(row.get("company_candidates", "")).strip()

    missing_core = not city or not district or not company
    info_score = fallback_info_score(row)

    if FALLBACK_REQUIRE_MISSING_CORE_FIELD and not missing_core:
        return False

    if info_score < FALLBACK_INFO_SCORE_THRESHOLD:
        return True

    if summary_len < MIN_SUMMARY_LENGTH_FOR_NO_FETCH and missing_core:
        return True

    if raw_len < MIN_ARTICLE_TEXT_LENGTH and missing_core:
        return True

    return False
async def enrich_articles_with_browser(df):
    """
    v20: Eksik bilgi skoru ile seçilen haberlerde browser fallback çalışır.
    Seri tek page yerine sınırlı paralel page kullanılır; böylece hızlı internet varken bekleme ciddi azalır.
    """
    if df.empty or not USE_BROWSER_ARTICLE_FETCH_FALLBACK:
        if "resolved_article_url" not in df.columns:
            df["resolved_article_url"] = ""
        return df

    df = df.copy()

    if "article_fetch_status" not in df.columns:
        df["article_fetch_status"] = "summary_only"
    if "resolved_article_url" not in df.columns:
        df["resolved_article_url"] = ""

    scored_candidates = []
    for idx, row in df.iterrows():
        if needs_browser_fallback(row):
            scored_candidates.append((fallback_priority_score(row), idx))

    scored_candidates.sort(reverse=True)
    candidate_indices = [idx for _, idx in scored_candidates[:BROWSER_ARTICLE_FETCH_MAX_ROWS]]

    if scored_candidates:
        print("Browser fallback aday toplamı:", len(scored_candidates))

    if not candidate_indices:
        return df

    print(f"\nBrowser içerik fallback başlıyor. Aday haber: {len(candidate_indices)} | concurrency={BROWSER_ARTICLE_CONCURRENCY}")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-blink-features=AutomationControlled",
                "--disable-extensions",
                "--disable-gpu"
            ]
        )

        sem = asyncio.Semaphore(BROWSER_ARTICLE_CONCURRENCY)
        started_at = time.time()

        async def worker(pos, idx):
            if time.time() - started_at > BROWSER_ARTICLE_TOTAL_TIME_LIMIT_SECONDS:
                return idx, "", "", "browser_total_timeout"
            async with sem:
                page = None
                try:
                    page = await browser.new_page(
                        viewport={"width": 1280, "height": 720},
                        user_agent=HEADERS["User-Agent"],
                        locale="tr-TR"
                    )
                    link = str(df.at[idx, "link"])
                    print(f"İçerik fallback {pos}/{len(candidate_indices)}")
                    return (idx, *await asyncio.wait_for(
                        fetch_article_with_browser(page, link),
                        timeout=BROWSER_ARTICLE_SINGLE_TIMEOUT_SECONDS
                    ))
                except asyncio.TimeoutError:
                    return idx, "", "", "browser_timeout"
                except Exception:
                    return idx, "", "", "browser_error"
                finally:
                    if page is not None:
                        try:
                            await page.close()
                        except Exception:
                            pass

        tasks = [worker(n, idx) for n, idx in enumerate(candidate_indices, start=1)]
        for coro in asyncio.as_completed(tasks):
            idx, article_text, resolved_url, status = await coro
            if resolved_url:
                df.at[idx, "resolved_article_url"] = resolved_url
            if article_text:
                df.at[idx, "raw_text"] = article_text
                df.at[idx, "entity_source_text"] = build_entity_source_text(
                    df.at[idx, "title"], df.at[idx, "summary"], article_text
                )
                df.at[idx, "article_fetch_status"] = status
            else:
                if "entity_source_text" not in df.columns or not str(df.at[idx, "entity_source_text"]).strip():
                    df.at[idx, "entity_source_text"] = build_entity_source_text(
                        df.at[idx, "title"], df.at[idx, "summary"], df.at[idx, "raw_text"]
                    )
                df.at[idx, "article_fetch_status"] = status

        await browser.close()

    return df


def get_date_window():
    """
    TARGET_DATE doluysa sadece o günü tarar.
    TARGET_DATE None ise son DAYS_BACK gün çalışır.
    """
    if TARGET_DATE:
        start_dt = datetime.strptime(TARGET_DATE, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        end_dt = start_dt + timedelta(days=1)
    else:
        end_dt = datetime.now(timezone.utc)
        start_dt = end_dt - timedelta(days=DAYS_BACK)

    return start_dt, end_dt


def rss_url(query):
    if TARGET_DATE:
        start_dt, end_dt = get_date_window()
        after_date = start_dt.strftime("%Y-%m-%d")
        before_date = end_dt.strftime("%Y-%m-%d")
        q = quote(f"{query} after:{after_date} before:{before_date}")
    else:
        q = quote(f"{query} when:{DAYS_BACK}d")

    return (
        "https://news.google.com/rss/search?"
        f"q={q}&hl={LANG}&gl={COUNTRY}&ceid={COUNTRY}:{LANG}&scoring=n"
    )


def detect_category(title, summary):
    text = normalize_text(f"{title} {summary}")

    rules = {
        "Fiziksel Risk / Yangın": ["yangin", "fabrika yangini", "depo yangini", "tesis yangini"],
        "Kara Para / MASAK / Bahis": ["kara para", "para aklama", "masak", "bahis", "yasa disi bahis", "yasadisi bahis"],
        "Dolandırıcılık / Sahtecilik": ["dolandiricilik", "sahtecilik", "sahte fatura"],
        "İflas / Konkordato / Mali Sorun": ["iflas", "konkordato", "mali kriz", "borc krizi", "odeme guclugu", "tasfiye"],
        "Operasyon / Soruşturma / Suç": ["sorusturma", "gozalti", "tutuklama", "suc orgutu", "orgutlu suc"],
        "Banka / Finansal Kuruluş": ["banka", "odeme kurulusu", "elektronik para", "supheli islem", "para transferi"],
    }

    for category, words in rules.items():
        if any(w in text for w in words):
            return category

    return "Diğer"


def detect_activity_type(text):
    text_norm = normalize_text(text)

    rules = {
        "üretim tesisi": ["uretim tesisi", "imalat", "imalathane", "uretim"],
        "fabrika": ["fabrika", "fabrikasi"],
        "depo": ["depo", "deposu", "lojistik deposu", "antrepo"],
        "atölye": ["atolye", "atolyesi"],
        "şantiye": ["santiye", "santiyesi"],
        "akaryakıt tesisi": ["akaryakit", "istasyon", "dolum tesisi"],
        "finansal kuruluş": ["banka", "odeme kurulusu", "elektronik para kurulusu", "finans kurulus"],
        "şirket / firma": ["sirket", "firma", "limited", "anonim sirket", "holding"],
    }

    found = []
    for activity, words in rules.items():
        if any(w in text_norm for w in words):
            found.append(activity)

    return " | ".join(found)


def process_entry(term, entry, start_date, end_date):
    title = entry.get("title", "")
    clean_title_for_output = remove_source_from_title(title)
    link = entry.get("link", "")
    summary_html = entry.get("summary", "")
    summary_text = clean_spaces(BeautifulSoup(summary_html, "html.parser").get_text(" "))

    exclude_text = normalize_text(f"{title} {summary_text}")
    if any(normalize_text(x) in exclude_text for x in EXCLUDE_NEWS_KEYWORDS):
        return None

    published_dt = parse_google_news_date(entry)

    if published_dt is None:
        return None

    if published_dt < start_date or published_dt >= end_date:
        return None

    if not keyword_match(title, TITLE_KEYWORDS):
        return None

    should_fetch_article = (
        ENABLE_ARTICLE_FETCH
        or (FETCH_ARTICLE_IF_SUMMARY_SHORT and len(summary_text) < MIN_SUMMARY_LENGTH_FOR_NO_FETCH)
    )

    article_text, resolved_article_url = fetch_article(link, force=should_fetch_article) if should_fetch_article else ("", "")
    text_for_analysis = article_text if article_text else summary_text

    combined_text = f"{title} {summary_text} {text_for_analysis}"
    normalized_combined_text = normalize_text(combined_text)

    foreign_location_keywords = [
        "abd", "amerika", "usa", "united states", "avrupa", "asya", "afrika",
        "almanya", "fransa", "italya", "ispanya", "ingiltere", "uk",
        "hollanda", "belcika", "iran", "irak", "suriye", "israil",
        "cin", "japonya", "hindistan", "pakistan", "kanada", "meksika",
        "brezilya", "new york", "los angeles", "londra", "paris", "berlin",
        "peru", "çad", "cad"
    ]

    important_foreign_financial_keywords = [
        "iflas", "konkordato", "tasfiye", "borc krizi", "mali kriz",
        "lisans iptali", "mal varligi", "el konuldu",
        "kara para", "para aklama", "masak", "yaptirim", "ambargo",
        "tedarik krizi", "banka", "finans kurulusu", "odeme kurulusu",
        "havayolu", "hava yolu", "kripto borsa"
    ]

    is_foreign_news = any(normalize_text(k) in normalized_combined_text for k in foreign_location_keywords)
    has_important_foreign = any(normalize_text(k) in normalized_combined_text for k in important_foreign_financial_keywords)

    if is_foreign_news and not has_important_foreign:
        return None

    matched_content_words = [
        word for word in CONTENT_KEYWORDS
        if normalize_text(word) in normalized_combined_text
    ]

    if re.search(r"\b(a\.?\s*ş\.?|anonim\s+şirket)\b", combined_text, flags=re.IGNORECASE):
        matched_content_words.append("A.Ş.")

    if re.search(r"\b(ltd\.?\s*şti\.?|limited\s+şirket)\b", combined_text, flags=re.IGNORECASE):
        matched_content_words.append("Ltd. Şti.")

    if not matched_content_words:
        return None

    category = detect_category(title, summary_text)
    activity_type = detect_activity_type(combined_text)

    return {
        "published": published_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "search_term": term,
        "category": category,
        "title": clean_title_for_output,
        "summary": summary_text,
        "link": link,
        "matched_words": ", ".join(sorted(set(matched_content_words))),
        "activity_type": activity_type,
        "raw_text": text_for_analysis,
        "entity_source_text": build_entity_source_text(clean_title_for_output, summary_text, text_for_analysis),
        "article_fetch_status": "requests_full_article" if article_text else "summary_only",
        "resolved_article_url": resolved_article_url,
    }


def collect_term_news(term, start_date, end_date):
    rows = []

    try:
        feed = feedparser.parse(rss_url(term))

        print("\n==============================")
        print("Aranan kelime:", term)
        print("Gelen haber:", len(feed.entries))
        print("==============================")

        for entry in feed.entries[:MAX_NEWS]:
            row = process_entry(term, entry, start_date, end_date)
            if row:
                rows.append(row)

    except Exception as e:
        print("Term hata:", term, str(e))

    return rows


def collect_news():
    start_date, end_date = get_date_window()
    print("Taranan tarih aralığı:", start_date.strftime("%Y-%m-%d"), "->", end_date.strftime("%Y-%m-%d"))

    all_rows = []

    with ThreadPoolExecutor(max_workers=RSS_WORKERS) as executor:
        futures = {
            executor.submit(collect_term_news, term, start_date, end_date): term
            for term in SEARCH_TERMS
        }

        for future in as_completed(futures):
            rows = future.result()
            all_rows.extend(rows)

    df = pd.DataFrame(all_rows)

    if df.empty:
        df = pd.DataFrame(columns=[
            "published", "search_term", "category", "title", "summary",
            "link", "matched_words", "activity_type", "raw_text", "entity_source_text", "article_fetch_status", "resolved_article_url"
        ])
    else:
        df = df.drop_duplicates(subset=["link"]).reset_index(drop=True)

    print("\nToplam bulunan haber:", len(df))
    return df


# ============================================================
# ENTITY EXTRACTION
# ============================================================

def clean_candidate(text):
    text = clean_spaces(text)
    text = text.strip(" -–|.,:;()[]{}")
    text = re.sub(r"^\s*(son dakika|haber|haberleri|ekonomi haberleri)\s*[:\-–]?\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+[-–|]\s+(haber|haberler|son dakika|ekonomi|gzt|haber 7|emlak kulisi|ege telgraf|yeni şafak|habertürk).*$", "", text, flags=re.IGNORECASE)
    # Şirket eklerinden sonra gelen Türkçe hâl eklerini temizle: İAR AŞ'ye -> İAR AŞ
    text = re.sub(r"((?:A\.?\s*Ş\.?|AŞ|A\.?S\.?|Ltd\.?\s*Şti\.?|LTD\.?\s*ŞTİ\.?|Ltd\.?|LTD\.?)\s*)['’](?:ye|ya|yi|yı|yu|yü|nin|nın|nun|nün|de|da|den|dan|na|ne)$", r"\1", text, flags=re.IGNORECASE)
    return text.strip(" -–|.,:;()[]{}")


def has_bad_source_word(text):
    norm = normalize_text(text)
    return any(normalize_text(w) in norm for w in NEWS_SOURCE_WORDS)




def strip_byline_noise(text):
    """DHA/İHA gibi haberlerdeki muhabir adı + şehir byline parçalarını entity metninden temizler."""
    text = str(text or "")
    # Örnek: "Mert ORDU/İSTANBUL,(DHA)-" -> "İSTANBUL,(DHA)-" veya tamamen kaldır
    text = re.sub(r"\b[A-ZÇĞİÖŞÜ][a-zçğıöşü]+\s+[A-ZÇĞİÖŞÜ]{3,}\s*/\s*[A-ZÇĞİÖŞÜ]{3,}\s*,?\s*\(?DHA\)?\s*-?", " ", text)
    text = re.sub(r"\b[A-ZÇĞİÖŞÜ][a-zçğıöşü]+\s+[A-ZÇĞİÖŞÜ]{3,}\s*/\s*[A-ZÇĞİÖŞÜ]{3,}\s*,?\s*\(?İHA\)?\s*-?", " ", text)
    text = re.sub(r"\b[A-ZÇĞİÖŞÜ][a-zçğıöşü]+\s+[A-ZÇĞİÖŞÜ]{3,}\s*/\s*[A-ZÇĞİÖŞÜ]{3,}\s*,?\s*\(?AA\)?\s*-?", " ", text)
    return clean_spaces(text)


def has_city_context(original_text, city):
    """Şehir adı yalnız başına değil, lokasyon bağlamıyla geçiyorsa kabul edilir."""
    text = str(original_text or "")
    city_re = re.escape(str(city))
    norm = normalize_text(text)
    city_norm = normalize_text(city)

    # Kayseri’de, Balıkesir'in, Adana merkezli, Adana Cumhuriyet Başsavcılığı vb.
    patterns_raw = [
        rf"\b{city_re}\s*['’]?\s*(da|de|ta|te|dan|den|tan|ten|nin|nın|nun|nün|in|ın|un|ün)\b",
        rf"\b{city_re}\s+(merkezli|merkezinde|ilinde|ilinin|ilçesinde|ilcesinde|cumhuriyet\s+başsavcılığı|cumhuriyet\s+başsavcılığınca|cumhuriyet\s+bassavciligi|cumhuriyet\s+bassavciliginca)\b",
        rf"\b{city_re}\s*/\s*[A-ZÇĞİÖŞÜa-zçğıöşü]+\b",
    ]
    if any(re.search(p, text, flags=re.IGNORECASE) for p in patterns_raw):
        return True

    patterns_norm = [
        rf"\b{re.escape(city_norm)}\s+(merkezli|merkezinde|ilinde|ilinin|cumhuriyet bassavciligi|cumhuriyet bassavciliginca)\b",
    ]
    return any(re.search(p, norm) for p in patterns_norm)


def has_district_explicit_context_any(original_text, district):
    """İlçe adının gerçekten lokasyon olarak geçtiğini gösteren genel bağlam."""
    text = str(original_text or "")
    d = re.escape(str(district))
    patterns = [
        rf"\b{d}\s*['’]?\s*(da|de|ta|te|dan|den|tan|ten|nin|nın|nun|nün|in|ın|un|ün)\b",
        rf"\b{d}\s+(ilçesi|ilcesi|ilçesinde|ilcesinde|ilçesine|ilcesine|ilçeden|ilceden|ilçe|ilce)\b",
        rf"\b{d}\s+(osb|organize sanayi|sanayi sitesi|mahallesi|caddesi|sokağı|sokagi|kara yolu|karayolu)\b",
    ]
    # "Bodrum katı" gibi yanlış pozitifleri özel olarak ele
    dn = normalize_text(district)
    norm = normalize_text(text)
    if dn == "bodrum" and re.search(r"\bbodrum\s+(kat|kati|katinda|katindaki)\b", norm):
        return False
    return any(re.search(p, text, flags=re.IGNORECASE) for p in patterns)

def extract_city_candidates(text):
    text = strip_byline_noise(text)
    text_norm = normalize_text(text)
    found = []

    # 1) İl adı yalnızca lokasyon bağlamıyla geçerse doğrudan kabul edilir.
    for city in TURKEY_CITIES:
        if re.search(rf"\b{re.escape(normalize_text(city))}\b", text_norm) and has_city_context(text, city):
            found.append(city)

    # 2) İlçe/mahalle üzerinden il çıkarımı. İlçe için de açık lokasyon bağlamı gerekir.
    for city, data in LOCATION_DICTIONARY.items():
        for district in data.get("districts", []):
            if district_match_is_safe(text, text_norm, city, district):
                found.append(city)

        for neighborhood in data.get("neighborhoods", []):
            n_norm = normalize_text(neighborhood)
            if re.search(rf"\b{re.escape(n_norm)}\b", text_norm):
                found.append(city)

    return list(dict.fromkeys(found))[:5]

def is_valid_entity_candidate(candidate, max_words=5):
    candidate = clean_candidate(candidate)
    norm = normalize_text(candidate)

    if not candidate or len(candidate) < 3:
        return False

    if len(candidate.split()) > max_words:
        return False

    if has_bad_source_word(candidate):
        return False

    if any(bad in norm for bad in ENTITY_EXTRACTION_BLACKLIST):
        return False

    if any(bad in norm for bad in BAD_CANDIDATE_PHRASES):
        return False

    # Adayın neredeyse tamamı rakam/noktalama olmasın
    if len(re.sub(r"[^A-Za-zÇĞİÖŞÜçğıöşü]", "", candidate)) < 3:
        return False

    return True


def is_valid_district_candidate(candidate):
    candidate = clean_candidate(candidate)
    norm = normalize_text(candidate)

    if not is_valid_entity_candidate(candidate, max_words=2):
        return False

    if norm in [normalize_text(x) for x in TURKEY_CITIES]:
        return False

    if any(bad in norm for bad in DISTRICT_BLACKLIST):
        return False

    # İlçe adayı sektör/olay kelimesi gibi görünüyorsa ele
    blocked_tokens = [
        "yangin", "fabrika", "depo", "tesis", "mobilya", "otomotiv", "tekstil",
        "dava", "durusma", "haber", "ajans", "son", "dakika", "kritik"
    ]
    if any(tok in norm.split() for tok in blocked_tokens):
        return False

    return True


def extract_district_candidates(text):
    """
    v20:
    - Serbest fallback yok.
    - Sözlük eşleşmesi de tek başına yetmez.
    - İlçe için açık lokasyon bağlamı gerekir: Maltepe'de, Burhaniye ilçesinde, Balıkesir Burhaniye vb.
    """
    text = strip_byline_noise(text)
    text_norm = normalize_text(text)
    found = []

    for city, data in LOCATION_DICTIONARY.items():
        for district in data.get("districts", []):
            if district_match_is_safe(text, text_norm, city, district):
                found.append(district)

    return list(dict.fromkeys(found))[:5]

def extract_neighborhood_candidates(text):
    text_norm = normalize_text(text)
    found = []

    for city, data in LOCATION_DICTIONARY.items():
        for neighborhood in data.get("neighborhoods", []):
            if re.search(rf"\b{re.escape(normalize_text(neighborhood))}\b", text_norm):
                found.append(neighborhood)

    patterns = [
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+Mahallesi)\b",
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+OSB)\b",
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+Organize Sanayi)\b",
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+Sanayi Sitesi)\b",
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+Küçük Sanayi)\b",
    ]

    for p in patterns:
        for m in re.findall(p, str(text), flags=re.IGNORECASE):
            m = clean_candidate(m)
            if m and not has_bad_source_word(m):
                found.append(m)

    return list(dict.fromkeys(found))[:5]


def extract_street_candidates(text):
    patterns = [
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+Caddesi)\b",
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+Sokağı)\b",
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+Bulvarı)\b",
    ]

    found = []

    for p in patterns:
        for m in re.findall(p, str(text), flags=re.IGNORECASE):
            m = clean_candidate(m)
            if m and not has_bad_source_word(m):
                found.append(m)

    return list(dict.fromkeys(found))[:5]


def normalize_facility_candidate(candidate):
    candidate = clean_candidate(candidate)
    words = candidate.split()

    # Çok uzun adaylar genelde başlık cümlesinden kopuyor: "Davasında Kritik Duruşma Mobilya Fabrikası"
    # Sondaki tesis ifadesini koruyup baştaki haber/gürültü kelimelerini kesiyoruz.
    if len(words) > 4:
        candidate_norm = normalize_text(candidate)
        for suffix in FACILITY_SUFFIXES:
            suffix_words = suffix.split()
            suffix_len = len(suffix_words)
            if candidate_norm.endswith(normalize_text(suffix)):
                keep_len = min(4, suffix_len + 2)
                candidate = " ".join(words[-keep_len:])
                break

    return clean_candidate(candidate)


def valid_facility_candidate(candidate):
    candidate = normalize_facility_candidate(candidate)
    norm = normalize_text(candidate)
    words = candidate.split()

    if not is_valid_entity_candidate(candidate, max_words=5):
        return False

    if len(words) < 2:
        return False

    if len(words) > 5:
        return False

    if any(bad in norm for bad in ENTITY_EXTRACTION_BLACKLIST):
        return False

    # En az bir tesis/fiziksel yapı kelimesi olmalı
    required = [
        "fabrika", "fabrikasi", "depo", "deposu", "tesis", "tesisi", "atolye",
        "atolyesi", "akaryakit", "istasyon", "santiye", "uretim", "lojistik", "dagitim"
    ]
    if not any(r in norm for r in required):
        return False

    return True


def extract_facility_candidates(text):
    text = str(text)
    found = []

    patterns = [
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+fabrikası)\b",
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+deposu)\b",
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+tesisi)\b",
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+atölyesi)\b",
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+akaryakıt istasyonu)\b",
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+şantiyesi)\b",
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+üretim tesisi)\b",
        r"\b([A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}(?:\s+[A-ZÇĞİÖŞÜa-zçğıöşü0-9]{2,30}){0,3}\s+dağıtım merkezi)\b",
    ]

    for p in patterns:
        for m in re.findall(p, text, flags=re.IGNORECASE):
            m = normalize_facility_candidate(m)
            if valid_facility_candidate(m):
                found.append(m)

    return list(dict.fromkeys(found))[:5]

def valid_company_candidate(candidate):
    candidate = clean_candidate(candidate)

    if not candidate:
        return False

    norm = normalize_text(candidate)
    words = candidate.split()

    # TL'Yİ AŞAN gibi ifadelerden kırpılan "Yİ AŞ" sahte şirketlerini ele.
    if company_has_strong_legal_suffix(candidate):
        base_for_suffix = re.sub(r"\b(a\.?\s*ş\.?|a\.?\s*s\.?|aş|as|ltd\.?\s*şti\.?|ltd\.?|limited\s+şirketi|anonim\s+şirketi)\b.*$", "", candidate, flags=re.IGNORECASE).strip()
        base_norm_chars = re.sub(r"\s+", "", normalize_text(base_for_suffix))
        if len(base_norm_chars) < 3:
            return False

    # v14: kamu kurumu, kolluk, haber/jenerik finansal terim ve meslek adı firma değildir.
    if any(normalize_text(w) in norm for w in COMPANY_HARD_BLOCK_WORDS):
        return False

    if norm in [normalize_text(x) for x in COMPANY_GENERIC_PHRASES]:
        return False

    # Tek başına sektör/meslek gibi duran adayları alma: "Elektrik", "Elektronik", "Jandarma" vb.
    if len(words) <= 2 and not company_has_strong_legal_suffix(candidate):
        generic_short = {
            "elektrik", "elektronik", "elektronik para", "odeme kurulusu", "ödeme kuruluşu",
            "finansal kurulus", "finansal kuruluş", "jandarma", "emniyet", "polis",
            "firma", "sirket", "şirket", "isletme", "işletme", "kurulus", "kuruluş"
        }
        if norm in generic_short:
            return False

    if len(words) < 2:
        return False

    if len(words) > 8:
        return False

    if len(candidate) > 90:
        return False

    if has_bad_source_word(candidate):
        return False

    if any(bad in norm for bad in BAD_CANDIDATE_PHRASES):
        return False

    if any(bad in norm for bad in ENTITY_EXTRACTION_BLACKLIST):
        return False

    weak_start = [
        "dev", "buyuk", "unlu", "yillik", "kuresel", "turk", "turkiye",
        "sektorde", "zincirleme", "son", "hangi", "neden", "nasil",
        "milyonluk", "milyarlik", "kisinin", "kişinin", "efsane"
    ]

    if any(norm.startswith(w + " ") for w in weak_start):
        return False

    # Gerçek şirket formatı veya sektör eki yoksa alma.
    if not re.search(COMMERCIAL_ENTITY_PATTERN, candidate, flags=re.IGNORECASE):
        return False

    # "kişinin işine son verildi Otomotiv" gibi cümle kırpıntılarını ele.
    sentence_noise = ["isine son", "işine son", "son verildi", "kimdir", "neden", "hangi", "artirmayla", "artırmayla", "satisa cikti", "satışa çıktı", "asliye", "mahkemesi", "icra dairesi"]
    if any(x in norm for x in sentence_noise):
        return False

    return True


def extract_company_candidates(text):
    text = str(text)
    found = []

    sector_pattern = "|".join([re.escape(x) for x in VALID_COMPANY_SECTORS])

    patterns = [
        r"\b([A-ZÇĞİÖŞÜ][A-ZÇĞİÖŞÜa-zçğıöşü0-9&\.\- ]{2,70}?\s+A\.?Ş\.?)\b",
        r"\b([A-ZÇĞİÖŞÜ][A-ZÇĞİÖŞÜa-zçğıöşü0-9&\.\- ]{2,70}?\s+LTD\.?\s*ŞTİ\.?)\b",
        r"\b([A-ZÇĞİÖŞÜ][A-ZÇĞİÖŞÜa-zçğıöşü0-9&\.\- ]{2,70}?\s+Limited Şirketi)\b",
        r"\b([A-ZÇĞİÖŞÜ][A-ZÇĞİÖŞÜa-zçğıöşü0-9&\.\- ]{2,70}?\s+Anonim Şirketi)\b",
        rf"\b([A-ZÇĞİÖŞÜ][A-ZÇĞİÖŞÜa-zçğıöşü0-9&\.\-]{{1,30}}(?:\s+[A-ZÇĞİÖŞÜ][A-ZÇĞİÖŞÜa-zçğıöşü0-9&\.\-]{{1,30}}){{0,3}}\s+(?:{sector_pattern}))\b",
    ]

    for p in patterns:
        for m in re.findall(p, text, flags=re.IGNORECASE):
            m = clean_candidate(m)
            if valid_company_candidate(m):
                found.append(m)

    return list(dict.fromkeys(found))[:10]


def canonical_company_key(candidate):
    """Aynı şirketin A.Ş./AŞ/nokta varyasyonlarını tekilleştirmek için normalize anahtar."""
    norm = normalize_text(candidate)
    norm = re.sub(r"\b(a\s*s|as|anonim\s+sirketi|anonim\s+sirket|ltd|sti|limited\s+sirketi|limited\s+sirket)\b", " ", norm)
    norm = re.sub(r"\s+", " ", norm).strip()
    return norm


def company_has_strong_legal_suffix(candidate):
    """Aday gerçek legal suffix taşıyor mu?
    v27 fix: Eski regex içindeki çıplak "as" ifadesi Plastik/Basın vb. kelimelerin içinde
    yanlışlıkla AŞ sinyali üretiyordu. Artık suffix token sınırlarıyla aranır.
    """
    text = str(candidate or "")
    return bool(re.search(
        r"(?<![A-Za-zÇĞİÖŞÜçğıöşü0-9])"
        r"(?:A\.?\s*Ş\.?|AŞ|A\.?\s*S\.?|AS|Ltd\.?\s*Şti\.?|LTD\.?\s*ŞTİ\.?|ltd\.?\s*şti\.?|Limited\s+Şirketi|LİMİTED\s+ŞİRKETİ|limited\s+şirketi|Anonim\s+Şirketi|ANONİM\s+ŞİRKETİ|anonim\s+şirketi)"
        r"(?![A-Za-zÇĞİÖŞÜçğıöşü0-9])",
        text,
        flags=0
    ))


def is_bad_company_context(text, start, end, window=45):
    """Adayın çevresindeki bağlam kolluk/haber/jenerik işlem ise ele."""
    left = str(text or "")[max(0, start - window):start]
    right = str(text or "")[end:min(len(str(text or "")), end + window)]
    ctx = normalize_text(left + " " + right)
    return any(normalize_text(w) in ctx for w in COMPANY_CONTEXT_BAD_WORDS)


def company_candidate_has_name_signal(candidate):
    """Legal suffix yoksa adayda marka/özel ad sinyali arar."""
    candidate = clean_candidate(candidate)
    norm = normalize_text(candidate)
    words = candidate.split()

    if company_has_strong_legal_suffix(candidate):
        return True

    # Aday yalnızca sektör/generic kelimelerden oluşuyorsa gerçek firma sayma.
    generic_tokens = {
        "elektrik", "elektronik", "para", "odeme", "ödeme", "kurulusu", "kuruluşu",
        "kurulus", "kuruluş", "firma", "sirket", "şirket", "isletme", "işletme",
        "sanayi", "ticaret", "tekstil", "gida", "gıda", "lojistik", "enerji",
        "otomotiv", "insaat", "inşaat", "fabrika", "depo", "tesis"
    }
    tokens = set(norm.split())
    if tokens and tokens.issubset(generic_tokens):
        return False

    # En az bir kelime sektör dışı özel ad gibi durmalı.
    for w in words:
        w_norm = normalize_text(w)
        if len(w_norm) >= 3 and w_norm not in generic_tokens:
            return True

    return False

def company_is_shouting_body_noise(candidate):
    """v23: Body'den gelen çok uzun tamamen büyük harfli kırpıntıları ele.
    Örn: BEREKET ... LTD. ŞTİSAHİB. Gültav/İAR gibi normal adayları etkilemez.
    """
    candidate = clean_candidate(candidate)
    letters = re.findall(r"[A-Za-zÇĞİÖŞÜçğıöşü]", candidate)
    if not letters:
        return False
    upper = sum(1 for ch in letters if ch.upper() == ch and ch.lower() != ch)
    ratio = upper / max(len(letters), 1)
    words = candidate.split()
    norm = normalize_text(candidate)
    # Uzun, tümü büyük, ticaret/ltd içerikli body kırpıntısı genelde ilan/listeden geliyor.
    if len(words) >= 5 and ratio >= 0.82:
        return True
    if "stisahib" in norm or "sti sahib" in norm or "şti sahib" in candidate.lower():
        return True
    return False


def extract_company_candidates_with_positions(text):
    """Firma adaylarını metin içindeki pozisyonlarıyla döndürür; v18 daha sıkı ve case-sensitive."""
    text = str(text or "")
    sector_pattern = "|".join([re.escape(x) for x in VALID_COMPANY_SECTORS])

    # v23: Legal suffix'ten sonra harf bitişi varsa alma. Örn: ŞTİSAHİB -> şirket değil.
    # Ama İAR AŞ'ye / Gültav Ltd. Şti.'nin gibi Türkçe ekleri yakala ve clean_candidate temizlesin.
    legal_pattern = r"\b((?:[A-ZÇĞİÖŞÜ][A-ZÇĞİÖŞÜa-zçğıöşü0-9&\.\-]*|[A-ZÇĞİÖŞÜ]{2,})(?:\s+(?:ve|Ve|VE|&|[A-ZÇĞİÖŞÜ][A-ZÇĞİÖŞÜa-zçğıöşü0-9&\.\-]*|[A-ZÇĞİÖŞÜ]{2,})){0,8}\s+(?:A\.?\s*Ş\.?|A\.?\s*S\.?|Ltd\.?\s*Şti\.?|LTD\.?\s*ŞTİ\.?|ltd\.?\s*şti\.?|Limited\s+Şirketi|LİMİTED\s+ŞİRKETİ|limited\s+şirketi|Anonim\s+Şirketi|ANONİM\s+ŞİRKETİ|anonim\s+şirketi)(?:[\.'’][A-Za-zÇĞİÖŞÜçğıöşü]+)?)(?![A-Za-zÇĞİÖŞÜçğıöşü])"
    sector_pattern_full = rf"\b((?:[A-ZÇĞİÖŞÜ][A-ZÇĞİÖŞÜa-zçğıöşü0-9&\.\-]*|[A-ZÇĞİÖŞÜ]{{2,}})(?:\s+(?:ve|Ve|VE|&|[A-ZÇĞİÖŞÜ][A-ZÇĞİÖŞÜa-zçğıöşü0-9&\.\-]*|[A-ZÇĞİÖŞÜ]{{2,}})){{0,3}}\s+(?:{sector_pattern}))\b"

    patterns = [legal_pattern, sector_pattern_full]

    found = []
    seen = set()

    for pattern in patterns:
        for m in re.finditer(pattern, text):
            candidate = clean_candidate(m.group(1))
            key = normalize_text(candidate)
            if key in seen:
                continue
            seen.add(key)

            if not valid_company_candidate(candidate):
                continue
            if not company_candidate_has_name_signal(candidate):
                continue
            # Legal suffix gerçek şirket sinyalidir; ama mahkeme/kolluk bağlamı hâlâ elenir.
            if is_bad_company_context(text, m.start(1), m.end(1)) and not company_has_strong_legal_suffix(candidate):
                continue
            if any(normalize_text(w) in key for w in ["asliye", "mahkemesi", "icra dairesi", "ticaret mahkemesi"]):
                continue

            found.append({"candidate": candidate, "start": m.start(1), "end": m.end(1)})

    return found

def score_company_candidate_occurrence(candidate, area, start_pos=0, text_len=0):
    """
    v13 alan bazlı skor:
    - Başlık en güçlü sinyal
    - Summary güçlü sinyal
    - Gövde tamamen kullanılmaya devam eder ama tek başına daha düşük puan getirir
    - İlk 1200 karakter gibi sert kesme yok; sadece pozisyona göre yumuşak ağırlık var
    """
    area = area or "body"

    if area == "title":
        score = COMPANY_TITLE_WEIGHT
    elif area == "summary":
        score = COMPANY_SUMMARY_WEIGHT
    else:
        score = COMPANY_BODY_WEIGHT
        if text_len and start_pos >= 0:
            ratio = start_pos / max(text_len, 1)
            if ratio <= 0.15:
                score += 18
            elif ratio <= 0.35:
                score += 10
            elif ratio <= 0.65:
                score += 5

    if company_has_strong_legal_suffix(candidate):
        score += 18
    else:
        score += 6

    word_count = len(str(candidate).split())
    if 2 <= word_count <= 5:
        score += 8
    elif word_count > 6:
        score -= 8

    return score



def extract_strict_legal_company_candidates(text):
    """Raw/entity metinden güçlü legal suffix taşıyan şirketleri yakalar.
    v25 kritik fix:
    - Metni cümleye bölmeden tarar; böylece "Ltd. Şti." noktadan parçalanıp kaybolmaz.
    - "Gültav Tavukçuluk Ltd. Şti. hakkında" gibi yapıları yakalar.
    - "İAR AŞ'ye" gibi ek almış AŞ yapılarını normalize eder.
    - "ŞTİSAHİB" gibi bitişik bozuk DOM kırpıntılarını almaz.
    """
    text = str(text or "")
    if not text.strip():
        return []

    # Çok gürültülü whitespace'i temizle ama noktaları koru.
    scan_text = clean_spaces(text)

    legal_regex = re.compile(
        r"\b("
        r"(?:[A-ZÇĞİÖŞÜ][A-Za-zÇĞİÖŞÜçğıöşü0-9&\-\.]{1,35}|[A-ZÇĞİÖŞÜ]{2,})"
        r"(?:\s+(?:ve|Ve|VE|&|[A-ZÇĞİÖŞÜ][A-Za-zÇĞİÖŞÜçğıöşü0-9&\-\.]{1,35}|[A-ZÇĞİÖŞÜ]{2,})){0,6}"
        r"\s+(?:A\.?\s*Ş\.?|AŞ|A\.?\s*S\.?|Ltd\.?\s*Şti\.?|LTD\.?\s*ŞTİ\.?|Limited\s+Şirketi|Anonim\s+Şirketi)"
        r")"
        r"(?=$|[\s,.;:!?)]|['’](?:ye|ya|yi|yı|yu|yü|nin|nın|nun|nün|de|da|den|dan|na|ne)\b)",
        flags=0,
    )

    found = []
    seen = set()
    for m in legal_regex.finditer(scan_text):
        cand = clean_candidate(m.group(1))
        key = normalize_text(cand)
        if not key or key in seen:
            continue

        # Suffix'in hemen devamında harf bitişi varsa bozuk DOM kırpıntısı olabilir: ŞTİSAHİB vb.
        after = scan_text[m.end(1):m.end(1) + 12]
        if re.match(r"^[A-Za-zÇĞİÖŞÜçğıöşü]", after):
            continue

        if company_is_shouting_body_noise(cand):
            continue
        if not valid_company_candidate(cand):
            continue

        # Kamu/kolluk/mahkeme bağlamında oluşan sahte legal adayları ele.
        sent_start = max(0, scan_text.rfind('.', 0, m.start(1)))
        sent_end = scan_text.find('.', m.end(1))
        if sent_end == -1:
            sent_end = min(len(scan_text), m.end(1) + 160)
        sent = scan_text[sent_start:sent_end]
        sent_norm = normalize_text(sent)
        if any(x in sent_norm for x in ["mahkeme", "savcilik", "savcılık", "jandarma", "emniyet", "polis", "icra dairesi"]):
            # Firma ismi mahkeme cümlesinde geçebilir; ama adayın kendisi mahkeme/kolluk değilse tamamen öldürme.
            if any(x in key for x in ["mahkeme", "savcilik", "jandarma", "emniyet", "polis", "icra"]):
                continue

        seen.add(key)
        found.append(cand)

    return found[:COMPANY_MAX_CANDIDATES]

def extract_company_candidates_scored(title, summary, raw_text):
    """
    Firma adaylarını başlık/summary/gövde alanlarına göre skorlar.
    Tam haber metni kaybedilmez; sadece gövdede tek başına çıkan zayıf adaylar elenir.
    """
    fields = [
        ("title", str(title or "")),
        ("summary", str(summary or "")),
        ("body", str(raw_text or "")),
    ]

    candidates = {}

    for area, text in fields:
        if not text.strip():
            continue

        for item in extract_company_candidates_with_positions(text):
            cand = item["candidate"]
            key = canonical_company_key(cand)
            if not key:
                continue

            occ_score = score_company_candidate_occurrence(cand, area, item["start"], len(text))

            if key not in candidates:
                candidates[key] = {
                    "display": cand,
                    "score": 0,
                    "areas": set(),
                    "count": 0,
                    "legal": company_has_strong_legal_suffix(cand),
                }

            candidates[key]["score"] += occ_score
            candidates[key]["areas"].add(area)
            candidates[key]["count"] += 1

            # v28: Legal şirketlerde en kısa kırpıntıyı değil, en tam ticari unvanı koru.
            # Örn: "Bereket Emlak" yerine "Bereket TB Emlak Bilişim Teknolojileri Ticaret Ltd. Şti." kalmalı.
            current_display = candidates[key]["display"]
            cand_legal = company_has_strong_legal_suffix(cand)
            current_legal = company_has_strong_legal_suffix(current_display)
            if cand_legal and not current_legal:
                candidates[key]["display"] = cand
            elif cand_legal and current_legal:
                # Aynı legal grupta daha uzun/tam olanı tercih et.
                if len(cand.split()) > len(current_display.split()) or len(cand) > len(current_display) + 8:
                    candidates[key]["display"] = cand
            elif cand_legal == current_legal and len(cand) < len(current_display):
                candidates[key]["display"] = cand

    scored = []
    for item in candidates.values():
        areas = item["areas"]
        score = item["score"]

        if len(areas) >= 2:
            score += 25
        if item["count"] >= 2:
            score += min(20, item["count"] * 4)

        # v16: Tam içerik açıkken gövdede tek başına geçen sektör benzeri adaylar çok gürültü üretir.
        # Ancak legal suffix taşıyan gerçek şirketleri (örn. "Gültav Tavukçuluk Ltd. Şti.") kaybetme.
        if areas == {"body"}:
            if item["legal"]:
                score += 12
            else:
                score -= 38
            # v23: Body-only gelen uzun tümü büyük ticaret sicili/listesi kırpıntılarını ele.
            if company_is_shouting_body_noise(item["display"]):
                score = min(score, COMPANY_MIN_SCORE - 1)

        # Title/summary görmeyen, legal suffix taşımayan adaylar Maps'e gitmesin.
        if areas == {"body"} and not item["legal"] and score < 95:
            score = min(score, COMPANY_MIN_SCORE - 1)

        scored.append({
            "candidate": item["display"],
            "score": round(score, 1),
            "areas": "+".join(sorted(areas)),
            "count": item["count"],
        })

    scored.sort(key=lambda x: x["score"], reverse=True)

    accepted = [x for x in scored if x["score"] >= COMPANY_MIN_SCORE]

    # v16: Hiç aday kalmadıysa güçlü legal suffix taşıyan body adayını da kabul et.
    # Bu, haber gövdesinde geçen "Gültav Tavukçuluk Ltd. Şti." gibi gerçek şirketleri kurtarır.
    if not accepted:
        accepted = [
            x for x in scored
            if (
                x["score"] >= 55
                and (
                    "title" in x["areas"]
                    or "summary" in x["areas"]
                    or company_has_strong_legal_suffix(x["candidate"])
                )
            )
        ]

    accepted = accepted[:COMPANY_MAX_CANDIDATES]

    preferred_candidates = prefer_full_legal_company_names([x["candidate"] for x in accepted])

    return (
        preferred_candidates,
        " | ".join([f"{x['candidate']}:{x['score']}:{x['areas']}" for x in accepted])
    )



def company_token_set_for_dedupe(candidate):
    """v28: Şirket adının legal suffix ve generic sektör kelimeleri dışındaki token seti."""
    norm = normalize_text(candidate)
    norm = re.sub(r"\b(a\s*s|as|a\s*s|anonim\s+sirketi|anonim\s+sirket|ltd|sti|limited\s+sirketi|limited\s+sirket)\b", " ", norm)
    generic = {
        "sanayi", "ticaret", "limited", "sirket", "sirketi", "anonim", "ltd", "sti",
        "teknolojileri", "teknoloji", "bilisim", "bilişim"
    }
    return {t for t in norm.split() if len(t) >= 2 and t not in generic}


def prefer_full_legal_company_names(candidates):
    """v28: Aynı şirket için kısa kırpıntı + tam unvan birlikte geldiyse tam legal unvanı bırakır.
    Örn: "Bereket Emlak" / "Bereket TB Emlak Bilişim Teknolojileri Ticaret Ltd. Şti." -> tam unvan.
    """
    cleaned = []
    for c in candidates or []:
        c = clean_candidate(c)
        if c and valid_company_candidate(c) and c not in cleaned:
            cleaned.append(c)

    keep = []
    for c in cleaned:
        c_tokens = company_token_set_for_dedupe(c)
        c_legal = company_has_strong_legal_suffix(c)
        drop = False
        for other in cleaned:
            if other == c:
                continue
            o_tokens = company_token_set_for_dedupe(other)
            o_legal = company_has_strong_legal_suffix(other)
            if not o_tokens or not c_tokens:
                continue
            # Tam legal unvan, kısa adayın ayırt edici tokenlarını kapsıyorsa kısa adayı at.
            if o_legal and len(other) > len(c) and c_tokens.issubset(o_tokens):
                drop = True
                break
            # İki legal adaydan tokenları kapsayan ve daha uzun olanı bırak.
            if c_legal and o_legal and len(other) > len(c) and c_tokens.issubset(o_tokens):
                drop = True
                break
        if not drop:
            keep.append(c)
    return keep[:COMPANY_MAX_CANDIDATES]

def join_clean_candidates(candidates, validator=None, limit=5):
    cleaned = []
    for candidate in candidates:
        candidate = clean_candidate(candidate)
        if not candidate:
            continue
        if validator and not validator(candidate):
            continue
        if candidate not in cleaned:
            cleaned.append(candidate)
    return " | ".join(cleaned[:limit])


def build_entity_quality_note(row):
    notes = []
    if not str(row.get("city_candidates", "")).strip():
        notes.append("il yok")
    if not str(row.get("district_candidates", "")).strip():
        notes.append("ilçe yok")
    if not str(row.get("company_candidates", "")).strip():
        notes.append("firma yok")
    if not str(row.get("facility_candidates", "")).strip():
        notes.append("faaliyet/tesis yok")
    return " | ".join(notes) if notes else "iyi"


def enrich_entities(df):
    rows = []

    for _, row in df.iterrows():
        title_text = str(row.get('title', ''))
        summary_text = str(row.get('summary', ''))
        raw_text = str(row.get('raw_text', ''))
        entity_text = str(row.get('entity_source_text', '')).strip()
        if not entity_text:
            entity_text = build_entity_source_text(title_text, summary_text, raw_text)
        full_text = entity_text

        city_candidates = extract_city_candidates(full_text)
        district_candidates = extract_district_candidates(full_text)
        neighborhood_candidates = extract_neighborhood_candidates(full_text)
        street_candidates = extract_street_candidates(full_text)
        facility_candidates = extract_facility_candidates(full_text)

        # v17: Firma adayları tam raw_text yerine sidebar temizlenmiş entity_source_text üzerinden skorlanır.
        # raw_text Excel'de tam kalır; ilgili haber/sidebar kaynaklı İmamoğlu/Çan/Merkez vb. sızıntılar azalır.
        company_candidates, company_candidate_scores = extract_company_candidates_scored(
            title_text, summary_text, entity_text
        )

        # v24: entity_source_text bazen Şti. kısmını kırpabiliyor. Bu yüzden raw_text içinden
        # sadece güçlü legal suffix taşıyan şirketleri güvenli şekilde geri ekliyoruz.
        raw_legal_companies = extract_strict_legal_company_candidates(raw_text)
        if raw_legal_companies:
            merged_companies = []
            for c in list(company_candidates) + raw_legal_companies:
                if valid_company_candidate(c) and c not in merged_companies:
                    merged_companies.append(c)
            company_candidates = prefer_full_legal_company_names(merged_companies)[:COMPANY_MAX_CANDIDATES]
            extra_scores = [f"{c}:raw_legal" for c in raw_legal_companies if c in company_candidates]
            if extra_scores:
                company_candidate_scores = (company_candidate_scores + " | " if company_candidate_scores else "") + " | ".join(extra_scores)

        enriched_row = {
            "city_candidates": join_clean_candidates(city_candidates, limit=5),
            "district_candidates": join_clean_candidates(district_candidates, limit=5),
            "neighborhood_candidates": join_clean_candidates(neighborhood_candidates, validator=lambda x: is_valid_entity_candidate(x, max_words=5), limit=5),
            "street_candidates": join_clean_candidates(street_candidates, validator=lambda x: is_valid_entity_candidate(x, max_words=5), limit=5),
            "facility_candidates": join_clean_candidates(facility_candidates, validator=valid_facility_candidate, limit=5),
            "company_candidates": join_clean_candidates(company_candidates, validator=valid_company_candidate, limit=COMPANY_MAX_CANDIDATES),
            "company_candidate_scores": company_candidate_scores,
        }
        enriched_row["entity_quality_note"] = build_entity_quality_note(enriched_row)
        rows.append(enriched_row)

    return pd.concat([df.reset_index(drop=True), pd.DataFrame(rows).reset_index(drop=True)], axis=1)


# ============================================================
# STRONG DEDUPLICATION
# ============================================================

def clean_text_for_dedup(text):
    text = normalize_text(text)

    for word in NEWS_SOURCE_WORDS:
        text = text.replace(normalize_text(word), " ")

    for word in GENERAL_NOISE_WORDS:
        text = re.sub(rf"\b{re.escape(normalize_text(word))}\b", " ", text)

    text = re.sub(r"\s+", " ", text).strip()
    return text


def first_pipe_value(value):
    parts = [x.strip() for x in str(value).split("|") if x.strip()]
    return parts[0] if parts else ""


def event_signature(row):
    parts = [
        str(row.get("category", "")),
        str(row.get("activity_type", "")),
        first_pipe_value(row.get("city_candidates", "")),
        first_pipe_value(row.get("district_candidates", "")),
        first_pipe_value(row.get("neighborhood_candidates", "")),
        first_pipe_value(row.get("facility_candidates", "")),
        first_pipe_value(row.get("company_candidates", "")),
        first_pipe_value(row.get("maps_best_guess", "")),
        str(row.get("title", "")),
    ]

    return clean_text_for_dedup(" ".join(parts))


def dedup_key_text(row):
    return clean_text_for_dedup(
        f"{row.get('category', '')} "
        f"{row.get('activity_type', '')} "
        f"{row.get('city_candidates', '')} "
        f"{row.get('district_candidates', '')} "
        f"{row.get('neighborhood_candidates', '')} "
        f"{row.get('facility_candidates', '')} "
        f"{row.get('company_candidates', '')} "
        f"{row.get('maps_best_guess', '')} "
        f"{row.get('title', '')} "
        f"{row.get('summary', '')}"
    )

def is_same_event(row1, row2):
    t1 = dedup_key_text(row1)
    t2 = dedup_key_text(row2)

    sig1 = event_signature(row1)
    sig2 = event_signature(row2)

    if not t1 or not t2:
        return False

    words1 = set(t1.split())
    words2 = set(t2.split())

    if not words1 or not words2:
        return False

    jaccard = len(words1 & words2) / len(words1 | words2)

    score_token = fuzz.token_set_ratio(t1, t2)
    score_partial = fuzz.partial_ratio(t1, t2)
    score_sort = fuzz.token_sort_ratio(t1, t2)

    title_score = fuzz.token_set_ratio(
        clean_text_for_dedup(row1.get("title", "")),
        clean_text_for_dedup(row2.get("title", ""))
    )

    sig_score = fuzz.token_set_ratio(sig1, sig2)

    same_category = str(row1.get("category", "")) == str(row2.get("category", ""))

    city1 = str(row1.get("city_candidates", "")).split("|")[0].strip()
    city2 = str(row2.get("city_candidates", "")).split("|")[0].strip()
    same_city = city1 and city2 and normalize_text(city1) == normalize_text(city2)

    district1 = str(row1.get("district_candidates", "")).split("|")[0].strip()
    district2 = str(row2.get("district_candidates", "")).split("|")[0].strip()
    same_district = district1 and district2 and normalize_text(district1) == normalize_text(district2)

    company1 = first_pipe_value(row1.get("company_candidates", ""))
    company2 = first_pipe_value(row2.get("company_candidates", ""))
    same_company = company1 and company2 and fuzz.token_set_ratio(normalize_text(company1), normalize_text(company2)) >= 88

    maps1 = first_pipe_value(row1.get("maps_best_guess", ""))
    maps2 = first_pipe_value(row2.get("maps_best_guess", ""))
    same_maps = maps1 and maps2 and fuzz.token_set_ratio(normalize_text(maps1), normalize_text(maps2)) >= 88

    # v7: Maps/firma aynıysa birleştirme sinyali güçlü ama yine kategori bağlamı aranıyor.
    # 65 gibi agresif genel eşik kullanılmıyor.
    if same_category and (same_company or same_maps) and score_token >= 70:
        return True

    if sig_score >= 90 and same_category:
        return True

    if same_category and same_city and score_token >= 72:
        return True

    if same_category and same_district and score_partial >= 78:
        return True

    if score_token >= 86:
        return True

    if same_category and score_token >= 76 and jaccard >= 0.30:
        return True

    if same_category and score_partial >= 88 and jaccard >= 0.28:
        return True

    if same_category and score_sort >= 80 and jaccard >= 0.32:
        return True

    if title_score >= 85:
        return True

    if same_category and jaccard >= 0.48:
        return True

    return False


def quality_score(row):
    """
    Duplicate grubu içinden hangi haberin kalacağını belirleyen skor.
    v7 mantığı: Dedup artık Maps/enrichment sonrasında çalıştığı için
    daha fazla bilgi içeren haber tercih edilir.
    """
    score = 0

    title = str(row.get("title", ""))
    summary = str(row.get("summary", ""))
    raw_text = str(row.get("raw_text", ""))

    city = str(row.get("city_candidates", "")).strip()
    district = str(row.get("district_candidates", "")).strip()
    neighborhood = str(row.get("neighborhood_candidates", "")).strip()
    street = str(row.get("street_candidates", "")).strip()
    facility = str(row.get("facility_candidates", "")).strip()
    company = str(row.get("company_candidates", "")).strip()
    maps_best = str(row.get("maps_best_guess", "")).strip()
    maps_candidates = str(row.get("maps_candidates", "")).strip()

    # Metin kalitesi
    score += len(title) * 0.5
    score += min(len(summary), 700) * 0.25
    score += min(len(raw_text), 1500) * 0.01

    if len(summary) > 80:
        score += 8
    if len(summary) > 180:
        score += 8

    if str(row.get("link", "")).startswith("http"):
        score += 8

    # Bilgi zenginliği: kontrolcü açısından en değerli alanlar
    if city:
        score += 10
    if district:
        score += 18
    if neighborhood:
        score += 22
    if street:
        score += 12
    if facility:
        score += 24
    if company:
        score += 34
        # v13: Firma adayı başlık/summary/gövde skoruyla güçlüyse duplicate seçiminde öne geçsin.
        score_text = str(row.get("company_candidate_scores", ""))
        score_values = []
        for part in score_text.split("|"):
            bits = part.strip().rsplit(":", 2)
            if len(bits) >= 2:
                try:
                    score_values.append(float(bits[-2]))
                except Exception:
                    pass
        if score_values:
            score += min(30, max(score_values) * 0.20)
    if maps_best:
        score += 30
    if maps_candidates:
        score += 12

    # Çok zayıf, clickbait/noise başlıkları geriye at
    noisy = any(normalize_text(x) in normalize_text(title) for x in NEWS_SOURCE_WORDS)
    if noisy:
        score -= 10

    weak_title_bits = ["hangi sirket", "devini", "iflasa surukledi", "son dakika"]
    if any(x in normalize_text(title) for x in weak_title_bits):
        score -= 15

    # v7: yanlış entity çıkarımları kalite skorunu yükseltmesin.
    bad_entity_text = normalize_text(" ".join([district, facility, company]))
    if any(bad in bad_entity_text for bad in ENTITY_EXTRACTION_BLACKLIST):
        score -= 25

    if str(row.get("entity_quality_note", "")).strip() == "iyi":
        score += 10

    return round(score, 2)


def add_record_quality_score(df):
    if df.empty:
        df["record_quality_score"] = []
        return df

    df = df.copy()
    df["record_quality_score"] = df.apply(quality_score, axis=1)
    return df


def deduplicate_news(df):
    if df.empty:
        return df

    df_work = df.copy().reset_index(drop=True)

    groups = []
    used = set()

    for i in range(len(df_work)):
        if i in used:
            continue

        group = [i]
        used.add(i)

        for j in range(i + 1, len(df_work)):
            if j in used:
                continue

            if is_same_event(df_work.loc[i], df_work.loc[j]):
                group.append(j)
                used.add(j)

        groups.append(group)

    keep_indices = []
    duplicate_group_ids = {}

    for group_id, group in enumerate(groups, start=1):
        best_idx = max(group, key=lambda idx: quality_score(df_work.loc[idx]))
        keep_indices.append(best_idx)

        for idx in group:
            duplicate_group_ids[idx] = group_id

    df_work["duplicate_group_id"] = [duplicate_group_ids.get(i, None) for i in range(len(df_work))]
    df_unique = df_work.loc[keep_indices].copy().reset_index(drop=True)

    print("\nTekilleştirme öncesi:", len(df))
    print("Tekilleştirme sonrası:", len(df_unique))
    print("Silinen benzer haber:", len(df) - len(df_unique))

    return df_unique


# ============================================================
# MAPS
# ============================================================



def meaningful_tokens_for_match(text):
    norm = normalize_text(text)
    stop = {
        "as", "a", "s", "ltd", "sti", "limited", "anonim", "sirket", "sirketi",
        "turkiye", "istanbul", "ankara", "izmir", "sanayi", "ticaret", "holding",
        "ve", "ile", "the", "google", "maps", "haritalar"
    }
    return [t for t in norm.split() if len(t) >= 3 and t not in stop]


def acronym_from_company_name(name):
    words = meaningful_tokens_for_match(name)
    if not words:
        return ""
    return "".join(w[0] for w in words[:5])


def maps_candidate_score(place, references):
    if not place or not references:
        return 0

    place_norm = normalize_text(place)
    best = 0

    for ref in references:
        ref_norm = normalize_text(ref)
        if not ref_norm:
            continue

        score = fuzz.token_set_ratio(place_norm, ref_norm)

        # Önemli kelimelerin kesişimi: İstanbul Altın Rafinerisi AŞ -> İstanbul Altın Rafinerisi
        ref_tokens = set(meaningful_tokens_for_match(ref))
        place_tokens = set(meaningful_tokens_for_match(place))
        if ref_tokens:
            overlap = len(ref_tokens & place_tokens) / len(ref_tokens)
            score = max(score, int(overlap * 100))

        # Kısaltma desteği: İstanbul Altın Rafinerisi -> IAR / İAR benzeri
        acronym = acronym_from_company_name(ref)
        if acronym and len(acronym) >= 3 and acronym in place_norm.replace(" ", ""):
            score = max(score, 85)

        best = max(best, score)

    return best


def filter_maps_candidates_for_row(row, places):
    """
    Maps ilk sonucu bazen alakasız işletme döndürüyor.
    Haber içinde firma adayı varsa Maps sonucu firma ile anlamlı benzerlik taşımak zorunda.
    Firma yoksa tesis/lokasyon aramasında daha gevşek davranılır.
    """
    companies = [x.strip() for x in str(row.get("company_candidates", "")).split("|") if x.strip()]
    facilities = [x.strip() for x in str(row.get("facility_candidates", "")).split("|") if x.strip()]

    filtered = []
    scores = []
    rejected = []

    if companies:
        references = companies
        min_score = 48
    elif facilities:
        references = facilities
        min_score = 35
    else:
        references = []
        min_score = 0

    for place in places:
        place = clean_candidate(place)
        if not place:
            continue
        if has_bad_source_word(place):
            rejected.append(place)
            continue

        if references:
            score = maps_candidate_score(place, references)
            if score < min_score:
                rejected.append(place)
                continue
        else:
            score = 0

        if place not in filtered:
            filtered.append(place)
            scores.append(str(score))

    return filtered, scores, rejected

def valid_maps_query(q):
    q_norm = normalize_text(q)

    if len(q) < 10:
        return False

    if any(bad in q_norm for bad in BAD_CANDIDATE_PHRASES):
        return False

    if q_norm.count("turkiye") > 1:
        return False

    weak_fragments = [
        "devini", "surukledi", "haber", "gazete",
        "son dakika", "de tekstil", "te tekstil",
        "daha iflasa", "hangi sirket"
    ]

    if any(w in q_norm for w in weak_fragments):
        return False

    if any(bad in q_norm for bad in ENTITY_EXTRACTION_BLACKLIST):
        return False

    return True


def build_maps_queries(row):
    companies = [x.strip() for x in str(row.get("company_candidates", "")).split("|") if x.strip()]
    facilities = [x.strip() for x in str(row.get("facility_candidates", "")).split("|") if x.strip()]
    cities = [x.strip() for x in str(row.get("city_candidates", "")).split("|") if x.strip()]
    districts = [x.strip() for x in str(row.get("district_candidates", "")).split("|") if x.strip()]
    neighborhoods = [x.strip() for x in str(row.get("neighborhood_candidates", "")).split("|") if x.strip()]
    streets = [x.strip() for x in str(row.get("street_candidates", "")).split("|") if x.strip()]

    queries = []

    location_parts = []
    if streets:
        location_parts.append(streets[0])
    if neighborhoods:
        location_parts.append(neighborhoods[0])
    if districts:
        location_parts.append(districts[0])
    if cities:
        location_parts.append(cities[0])

    location_text = " ".join(location_parts).strip()

    if companies:
        for company in companies[:2]:
            queries.append(f"{company} {location_text} Türkiye".strip())

    elif facilities and (districts or neighborhoods or streets):
        for facility in facilities[:1]:
            queries.append(f"{facility} {location_text} Türkiye".strip())

    else:
        return []

    final_queries = []
    seen = set()

    for q in queries:
        q = clean_spaces(q)
        q_norm = normalize_text(q)

        if not valid_maps_query(q):
            continue

        if q_norm not in seen:
            seen.add(q_norm)
            final_queries.append(q)

    return final_queries[:MAPS_MAX_QUERIES_PER_NEWS]


async def extract_place_names_from_maps_page(page, max_places=5):
    """
    Google Maps sonuçlarından işletme adlarını toplar.
    v11: Sponsorlu/Reklam kartları atlanır; böylece Agakulche gibi reklam sonuçları aday olmaz.
    """
    place_names = []
    selectors = ["div.Nv2PK", "div[role='article']", "a.hfpxzc"]

    for selector in selectors:
        try:
            elements = await page.query_selector_all(selector)

            for el in elements:
                card_text = ""

                try:
                    card_text = await el.inner_text(timeout=1000)
                except Exception:
                    card_text = ""

                # a.hfpxzc seçildiyse en yakın sonuç kartının metnini almaya çalış
                if not card_text or selector == "a.hfpxzc":
                    try:
                        parent_text = await el.evaluate("""
                            node => {
                                const card = node.closest('div.Nv2PK') || node.closest("div[role='article']") || node.parentElement;
                                return card ? card.innerText : '';
                            }
                        """)
                        if parent_text:
                            card_text = parent_text
                    except Exception:
                        pass

                card_text_norm = normalize_text(card_text)

                if any(normalize_text(w) in card_text_norm for w in SPONSORED_MAPS_WORDS):
                    continue

                name = ""

                try:
                    aria = await el.get_attribute("aria-label")
                    if aria:
                        name = aria.strip()
                except Exception:
                    pass

                if not name:
                    try:
                        txt = card_text or await el.inner_text(timeout=1000)
                        if txt:
                            # Kartta ilk satır çoğunlukla işletme adıdır.
                            lines = [clean_spaces(x) for x in txt.split("\n") if clean_spaces(x)]
                            if lines:
                                name = lines[0]
                    except Exception:
                        pass

                name = name.replace(" - Google Maps", "").replace(" - Google Haritalar", "").strip()
                name_norm = normalize_text(name)

                if not name:
                    continue

                if name_norm in [
                    "google maps", "google haritalar", "yol tarifi",
                    "kaydet", "arama", "paylas", "web sitesi", "rotalar", "baslat",
                    "sonuclar", "results"
                ]:
                    continue

                if has_bad_source_word(name):
                    continue

                if any(normalize_text(w) in name_norm for w in SPONSORED_MAPS_WORDS):
                    continue

                if 3 <= len(name) <= 100 and name not in place_names:
                    place_names.append(name)

                if len(place_names) >= max_places:
                    break

        except Exception:
            pass

        if place_names:
            break

    return place_names[:max_places]


async def maps_search_candidates(query, page, max_places=5):
    url = "https://www.google.com/maps/search/" + quote(query)

    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=25000)
        await page.wait_for_timeout(2500)

        place_names = await extract_place_names_from_maps_page(page, max_places=max_places)

        if not place_names:
            title = await page.title()
            title = title.replace(" - Google Maps", "").replace(" - Google Haritalar", "").strip()

            if title and 3 <= len(title) <= 100 and not has_bad_source_word(title):
                place_names = [title]

        return place_names

    except Exception:
        return []


async def enrich_maps(df):
    if df.empty or not USE_MAPS:
        df["maps_queries"] = ""
        df["maps_candidates"] = ""
        df["maps_best_guess"] = ""
        df["maps_status"] = "maps kapalı"
        df["maps_match_scores"] = ""
        df["maps_rejected_candidates"] = ""
        return df

    maps_queries_all = []
    maps_candidates_all = []
    maps_best_guess_all = []
    maps_status_all = []
    maps_match_scores_all = []
    maps_rejected_candidates_all = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-blink-features=AutomationControlled"
            ]
        )

        page = await browser.new_page(
            viewport={"width": 1366, "height": 768},
            user_agent=HEADERS["User-Agent"],
            locale="tr-TR"
        )

        for idx, row in df.iterrows():
            print(f"\nMaps {idx + 1}/{len(df)} işleniyor...")

            queries = build_maps_queries(row)
            candidates = []
            match_scores = []
            rejected_candidates = []

            if not queries:
                status = "maps araması yapılmadı"
                print(status)
            else:
                for q in queries:
                    print("Maps araması:", q)

                    found = await maps_search_candidates(q, page, max_places=MAPS_MAX_RESULTS)

                    filtered_found, filtered_scores, rejected_found = filter_maps_candidates_for_row(row, found)

                    for place, score in zip(filtered_found, filtered_scores):
                        if place not in candidates:
                            candidates.append(place)
                            match_scores.append(score)

                    for place in rejected_found:
                        if place not in rejected_candidates:
                            rejected_candidates.append(place)

                    await page.wait_for_timeout(800)

                    if len(candidates) >= MAPS_MAX_RESULTS:
                        break

                if candidates:
                    if str(row.get("company_candidates", "")).strip():
                        status = "firma adı + güvenli maps adayı var"
                    else:
                        status = "firma adı yok, tesis/lokasyon ile maps tahmini var"
                else:
                    if rejected_candidates:
                        status = "maps sonucu düşük güven nedeniyle elendi"
                    else:
                        status = "maps adayı yok"

            maps_queries_all.append(" | ".join(queries))
            maps_candidates_all.append(" | ".join(candidates))
            maps_best_guess_all.append(candidates[0] if candidates else "")
            maps_status_all.append(status)
            maps_match_scores_all.append(" | ".join(match_scores))
            maps_rejected_candidates_all.append(" | ".join(rejected_candidates))

        await browser.close()

    df = df.copy()
    df["maps_queries"] = maps_queries_all
    df["maps_candidates"] = maps_candidates_all
    df["maps_best_guess"] = maps_best_guess_all
    df["maps_status"] = maps_status_all
    df["maps_match_scores"] = maps_match_scores_all
    df["maps_rejected_candidates"] = maps_rejected_candidates_all

    return df


# ============================================================
# OUTPUT
# ============================================================

def export_csv(df, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    print("CSV hazır:", path)


def _write_sheet_from_df(wb, sheet_name, df, widths=None):
    ws = wb.create_sheet(sheet_name)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    headers = [cell.value for cell in ws[1]]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    widths = widths or {}
    for col_name, width in widths.items():
        if col_name in headers:
            idx = headers.index(col_name) + 1
            col_letter = ws.cell(row=1, column=idx).column_letter
            ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    if "link" in headers:
        link_col = headers.index("link") + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=link_col)
            if cell.value and str(cell.value).startswith("http"):
                url = cell.value
                cell.hyperlink = url
                cell.value = "Habere Git"
                cell.style = "Hyperlink"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    return ws


def export_excel(df, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    # Kontrolcü için sade ekran.
    # Teknik kolonlar, skorlar ve ara kolonlar sadece Raw_Data içinde kalır.
    review_columns = [
        "published",
        "category",
        "title",
        "summary",
        "link",
        "city_candidates",
        "district_candidates",
        "company_candidates",
        "maps_best_guess",
        "maps_status",
    ]

    available_review_columns = [c for c in review_columns if c in df.columns]
    review_df = df[available_review_columns].copy()

    review_widths = {
        "published": 20,
        "category": 30,
        "title": 70,
        "summary": 100,
        "link": 18,
        "city_candidates": 24,
        "district_candidates": 28,
        "company_candidates": 50,
        "maps_best_guess": 45,
        "maps_status": 35,
    }

    raw_widths = {
        "published": 20,
        "search_term": 25,
        "category": 30,
        "title": 60,
        "summary": 90,
        "link": 18,
        "matched_words": 40,
        "activity_type": 30,
        "city_candidates": 25,
        "district_candidates": 30,
        "neighborhood_candidates": 35,
        "street_candidates": 35,
        "facility_candidates": 40,
        "company_candidates": 50,
        "company_candidate_scores": 70,
        "maps_queries": 60,
        "maps_candidates": 60,
        "maps_best_guess": 40,
        "maps_status": 35,
        "entity_quality_note": 32,
        "duplicate_group_id": 18,
        "record_quality_score": 18,
        "article_fetch_status": 18,
        "entity_source_text": 100,
        "raw_text": 100,
    }

    _write_sheet_from_df(wb, "Kontrol_Listesi", review_df, review_widths)
    _write_sheet_from_df(wb, "Raw_Data", df, raw_widths)

    wb.save(path)
    print("Excel hazır:", path)



ENTITY_OUTPUT_COLUMNS = [
    "city_candidates", "district_candidates", "neighborhood_candidates", "street_candidates",
    "facility_candidates", "company_candidates", "company_candidate_scores", "entity_quality_note"
]


def refresh_entities(df):
    """
    v15: Browser fallback raw_text'i güncelledikten sonra entity alanlarını yeniden üretir.
    Aksi halde içerik çekilmiş olsa bile eski boş city/company alanları kalabilir.
    """
    df = df.copy()
    drop_cols = [c for c in ENTITY_OUTPUT_COLUMNS if c in df.columns]
    if drop_cols:
        df = df.drop(columns=drop_cols)
    return enrich_entities(df)
# ============================================================
# RUN
# ============================================================

print("Pipeline başladı.")

df = collect_news()

# v15: Önce summary/requests verisiyle entity çıkarılır.
# Sonra sadece city/district/company gibi kritik bilgisi eksik kalan değerli haberlerde browser fallback çalışır.
df = enrich_entities(df)
df = asyncio.run(enrich_articles_with_browser(df))

# v15: Fallback raw_text'i güncellediyse entity alanları yeniden hesaplanır.
df = refresh_entities(df)

# v8: Maps enrichment duplicate elemeden ÖNCE çalışır.
# Böylece aynı olay grubunda firma/lokasyon/maps bilgisi daha dolu olan haber korunur.
if USE_MAPS:
    print("\nMaps enrichment başlıyor.")
    df = asyncio.run(enrich_maps(df))
    print("Maps enrichment tamamlandı.")
else:
    df["maps_queries"] = ""
    df["maps_candidates"] = ""
    df["maps_best_guess"] = ""
    df["maps_status"] = "maps kapalı"
    df["maps_match_scores"] = ""
    df["maps_rejected_candidates"] = ""

# Dedup öncesi kalite skoru hesaplanır; dedup aynı grupta en zengin kaydı bırakır.
df = add_record_quality_score(df)
df = deduplicate_news(df)
df = add_record_quality_score(df)

export_csv(df, CSV_FILE)
export_excel(df, XLSX_FILE)

print("\nPipeline tamamlandı.")
print("Kayıt sayısı:", len(df))
